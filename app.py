from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash, abort
import sqlite3
import os
import io
from datetime import datetime, date, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.platypus import Image as RLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors as rl_colors

app = Flask(__name__)
app.secret_key = os.environ.get('CADUNICO_SECRET', 'cadunico2026secretkey_dev_only')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGOS_DIR = os.path.join(BASE_DIR, 'static', 'logos')

# ---------------------------------------------------------------------------
# Camada de banco de dados — SQLite (local) ou PostgreSQL (produção)
# ---------------------------------------------------------------------------
# Se a variável DATABASE_URL existir (Railway/Render), usa PostgreSQL.
# Caso contrário, usa SQLite local — zero configuração para desenvolvimento.

DATABASE_URL = os.environ.get('DATABASE_URL', '')
_DB_PATH = os.path.join(BASE_DIR, 'database.db')
_USE_PG = bool(DATABASE_URL)
_PG_FAILED = False

if _USE_PG:
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    try:
        import psycopg2
        import psycopg2.extras
    except ImportError:
        _USE_PG = False
        import sqlite3
else:
    import sqlite3


def _is_pg():
    return _USE_PG and not _PG_FAILED


class _Placeholder:
    def __str__(self):
        return '%s' if _is_pg() else '?'
    def __add__(self, other):
        return str(self) + str(other)
    def __radd__(self, other):
        return str(other) + str(self)

PH = _Placeholder()

# Fuso horário de Belém (UTC-3) — usado para determinar o ano da numeração VD
_TZ_BELEM = timezone(timedelta(hours=-3))


def _adapt_sql(sql: str) -> str:
    """Converte placeholders ? para %s quando usando PostgreSQL."""
    if _is_pg():
        return sql.replace('?', '%s')
    return sql


def get_db():
    """Retorna uma conexão com o banco configurado (PostgreSQL ou SQLite com fallback)."""
    global _USE_PG, _PG_FAILED
    if _USE_PG and not _PG_FAILED:
        try:
            conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
            return conn
        except Exception as e:
            app.logger.warning(f"[BANCO] Não foi possível conectar ao PostgreSQL ({e}). Alternando automaticamente para SQLite local ({_DB_PATH}).")
            _PG_FAILED = True

    import sqlite3
    conn = sqlite3.connect(_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _lastrowid(conn):
    """Retorna o ID do último INSERT (compatível com ambos os bancos)."""
    if _is_pg():
        return None  # PostgreSQL usa RETURNING na query
    else:
        return conn.cursor().lastrowid if hasattr(conn, 'cursor') else None


class _PGRow:
    """Wrapper para dict do psycopg2 se comportar como sqlite3.Row."""
    def __init__(self, d):
        self._d = d

    def __getitem__(self, key):
        return self._d[key]

    def __getattr__(self, key):
        try:
            return self._d[key]
        except KeyError:
            raise AttributeError(key)

    def get(self, key, default=None):
        return self._d.get(key, default)

    def keys(self):
        return self._d.keys()


def _exec(conn, sql, params=()):
    """Executa uma query adaptando o SQL e retornando cursor."""
    cur = conn.cursor()
    cur.execute(_adapt_sql(sql), params)
    return cur


def _fetchone(conn, sql, params=()):
    """Executa SELECT e retorna a primeira linha como Row-like."""
    cur = _exec(conn, sql, params)
    row = cur.fetchone()
    if row is None:
        return None
    if _is_pg():
        return _PGRow(dict(row))
    return row


def _fetchall(conn, sql, params=()):
    """Executa SELECT e retorna todas as linhas como lista de Row-like."""
    cur = _exec(conn, sql, params)
    rows = cur.fetchall()
    if _is_pg():
        return [_PGRow(dict(r)) for r in rows]
    return rows


def _logo_url(filename):
    """Retorna a URL da logo se o arquivo existir, senão None."""
    if os.path.exists(os.path.join(LOGOS_DIR, filename)):
        return url_for('static', filename=f'logos/{filename}')
    return None


@app.context_processor
def inject_logos():
    """Disponibiliza as logos em todos os templates via variáveis logo_*."""
    return dict(
        logo_prefeitura=_logo_url('prefeitura.png'),
        logo_setas=_logo_url('setas.png'),
        logo_cadunico=_logo_url('cadunico.png'),
        logo_bolsafamilia=_logo_url('bolsafamilia.png'),
    )

# ---------------------------------------------------------------------------
# Cloudinary — upload de anexos
# ---------------------------------------------------------------------------
import os as _os

_CLOUDINARY_URL = _os.environ.get('CLOUDINARY_URL', '')


def _upload_anexo(file_obj, pasta='visitas'):
    """Faz upload para Cloudinary e retorna (url, nome_original) ou (None, None)."""
    if not _CLOUDINARY_URL and not (
        _os.environ.get('CLOUDINARY_CLOUD_NAME') and
        _os.environ.get('CLOUDINARY_API_KEY') and
        _os.environ.get('CLOUDINARY_API_SECRET')
    ):
        return None, None
    try:
        import cloudinary
        import cloudinary.uploader
        cloudinary.config(
            cloud_name=_os.environ.get('CLOUDINARY_CLOUD_NAME', ''),
            api_key=_os.environ.get('CLOUDINARY_API_KEY', ''),
            api_secret=_os.environ.get('CLOUDINARY_API_SECRET', ''),
        )
        result = cloudinary.uploader.upload(
            file_obj,
            folder=pasta,
            resource_type='auto',
        )
        return result['secure_url'], file_obj.filename
    except Exception as e:
        app.logger.error(f'Cloudinary upload error: {e}')
        return None, None


TIPOS_ATENDIMENTO = [
    "Bloqueio de Benefício",
    "Desbloqueio de Benefício",
    "CadÚnico para BPC - 1ª Vez",
    "CadÚnico para BPC - Atualização",
    "Carteira do Idoso Emitida",
    "Comprovante de Cadastro",
    "Consulta Cadastro Único",
    "Consulta SIBEC",
    "Escuta Qualificada",
    "Benefício Eventual",
    "Atendimento Técnico / Parecer Social",
    "Exclusão de membros",
    "Folha de Pagamento (SIBEC)",
    "Inclusão de membros",
    "Inscrição do CadÚnico",
    "Recadastramento",
    "Reversão de Cancelamento",
    "Transferência de Município",
    "Tarifa Social de Energia Elétrica",
    "CadÚnico Unipessoal",
    "Visita Domiciliar",
    "Troca de RF",
]

TIPOS_SIBEC = {
    "Bloqueio de Benefício",
    "Desbloqueio de Benefício",
    "Consulta SIBEC",
    "Folha de Pagamento (SIBEC)",
    "Reversão de Cancelamento",
}
ORIGENS = ["Demanda Espontânea", "Encaminhado", "Visita Domiciliar"]

MOTIVOS_VISITA = [
    "Família Unipessoal",
    "Averiguação Cadastral",
    "Revisão Cadastral",
    "Busca Ativa",
    "Denúncia",
    "Suspeita de Inconsistência Cadastral",
    "Verificação de Endereço",
    "Verificação da Composição Familiar",
    "Inclusão Cadastral Domiciliar",
    "Atualização Cadastral Domiciliar",
    "Pessoa Idosa com dificuldade de locomoção",
    "Pessoa com Deficiência",
    "Pessoa Acamada",
    "Encaminhamento do CRAS",
    "Encaminhamento do CREAS",
    "Solicitação da Gestão do Cadastro Único",
    "Solicitação da Secretaria de Assistência Social",
    "Solicitação do Ministério Público",
    "Solicitação do Poder Judiciário",
    "Outro",
]

ORGAOS_ENCAMINHADORES = [
    "Demanda Espontânea (sem encaminhamento)",
    "CRAS",
    "CREAS",
    "SETAS",
    "SEMED",
    "SEMSA",
    "Conselho Tutelar – Sede",
    "Conselho Tutelar – 4 Bocas",
    "Ministério Público",
    "Defensoria Pública",
    "Poder Judiciário",
    "Equatorial Pará",
    "Outro",
]

MOTIVOS_ENCAMINHAMENTO = [
    "Inclusão no Cadastro Único",
    "Atualização Cadastral",
    "Averiguação Cadastral",
    "Revisão Cadastral",
    "Emissão de NIS",
    "Cadastro para BPC",
    "Programa Bolsa Família",
    "Tarifa Social de Energia",
    "Visita Domiciliar",
    "Busca Ativa",
    "Orientação",
    "Outro",
]

SITUACES_ENCAMINHAMENTO = ["Atendido", "Pendente", "Cancelado", "Não localizado"]


# ---------------------------------------------------------------------------
# Validação de CPF
# ---------------------------------------------------------------------------

def validar_cpf(cpf: str) -> bool:
    """Valida CPF pelos dígitos verificadores. Aceita somente dígitos."""
    cpf = ''.join(c for c in cpf if c.isdigit())
    if len(cpf) != 11 or len(set(cpf)) == 1:
        return False
    for pos in range(9, 11):
        soma = sum(int(cpf[i]) * (pos + 1 - i) for i in range(pos))
        digito = (soma * 10 % 11) % 10
        if digito != int(cpf[pos]):
            return False
    return True

# ---------------------------------------------------------------------------
# Banco de dados
# ---------------------------------------------------------------------------


def _config_defaults() -> dict:
    """Valores padrão das configurações editáveis do relatório."""
    return {
        'coordenadora': 'Rosiane Rodrigues Shiozaki',
        'setor_nome': 'Setor do Cadastro Único /PBF',
        'endereco': 'Bairro: Campina - Rua Bruno de Menezes s/n',
        'email_setor': 'setascadastrounico@gmail.com',
        'email_admin_notificacao': 'setascadastrounico@gmail.com',
        'smtp_host': 'smtp.gmail.com',
        'smtp_port': '587',
        'smtp_user': 'setascadastrounico@gmail.com',
        'smtp_pass': '',
        'territorio': 'Sede do Município e Distrito de Quatro Bocas',
        'texto_identificacao': (
            'Ao dia três de janeiro de 2024 (03.01.2024), iniciamos o ano letivo com realização de serviços '
            'internos elaboração de planejamentos das atividades, para que possamos ofertar um atendimento '
            'humanizado e de qualidade aos usuários do Sistema Único da Assistência Social - SUAS, no setor '
            'do Cadastro Único no Centro de Atendimento ao Cidadão - CAC em Tomé-Açu, no CRAS em Q. Bocas. '
            'O cronograma proposto consiste em incentivar as equipes das Unidades de Saúde Educação a '
            'realizarem parcerias "in loco", para que junto aos equipamentos sociais disponíveis, os '
            'profissionais possam desempenhar uma atividade intensiva de divulgação e esclarecimentos '
            'junto à comunidade.'
        ),
        'texto_apresentacao': (
            'A equipe do Cadastro Único é composta por sete (07) entrevistadores e dois (02) digitadores. '
            'No SIBEC Sistema de Benefício ao Cidadão, existe um Operador Comum treinado para operar esse '
            'sistema, o qual é gerenciado por um Usuário Master capacitado pela Caixa Econômica Federal '
            'para fazer a administração de Benefícios das famílias beneficiárias do Programa Bolsa Família '
            'no Município.\n\n'
            'Porém ao término de cada mês esse Departamento do Cadastro Único/PBF registra em planilha '
            'diárias todos os atendimentos realizados para serem inseridos no Sistema Mensal de Atendimentos '
            '(RMA), junto à equipe técnica do CRAS, com o objetivo de depurar os dados do Cad. Único para '
            'a qualidade dessas informações e melhoria do Cadastro Único e Programa Bolsa Família no Município.'
        ),
        'rodape': 'End: R. Pres. Costa e Silva - 68682-000 – Quatro Bocas - Tomé-Açu/PA\nFone: (91) 99254-0128 / Email: setascadastrounico@gmail.com',
        'municipio': 'Tomé-Açu',
        'prazo_visita_padrao': '15',
        'prazo_visita_Averiguacao Cadastral': '15',
        'prazo_visita_Revisao Cadastral': '20',
        'prazo_visita_BPC': '30',
        'prazo_visita_Denuncia': '7',
        'prazo_visita_Atualizacao Cadastral': '15',
        'prazo_visita_Inclusao Cadastral': '15',
        'visita_titulo_doc': 'SOLICITAÇÃO DE VISITA DOMICILIAR',
        'visita_subtitulo_doc': 'Secretaria Municipal de Assistência Social - Setor do Cadastro Único / PBF',
        'visita_orientacao_texto': (
            'A visita domiciliar é uma ação de acompanhamento cadastral e socioassistencial. '
            'O entrevistador/assistente social deve averiguar as informações declaradas, registrando o parecer '
            'técnico e garantindo a qualidade dos dados do Cadastro Único.'
        ),
        'visita_assinatura_1': 'Assinatura do Entrevistador / Assistente Social',
        'visita_assinatura_2': 'Assinatura do Responsável Familiar (RF)',
        'visita_rodape_txt': 'Setor do Cadastro Único e Programa Bolsa Família',
    }


def get_config() -> dict:
    """Retorna todas as configurações do relatório do banco."""
    conn = get_db()
    rows = _fetchall(conn, "SELECT chave, valor FROM config_relatorio")
    conn.close()
    cfg = _config_defaults()
    for row in rows:
        cfg[row['chave']] = row['valor']
    return cfg


def set_config(chave: str, valor: str):
    """Atualiza ou insere uma configuração."""
    conn = get_db()
    if _USE_PG:
        _exec(conn,
            "INSERT INTO config_relatorio (chave, valor) VALUES (%s, %s) ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
            (chave, valor)
        )
    else:
        _exec(conn,
            "INSERT OR REPLACE INTO config_relatorio (chave, valor) VALUES (?, ?)",
            (chave, valor)
        )
    conn.commit()
    conn.close()


def _processar_sla_visitas(visitas, cfg=None):
    """
    Calcula dias decorridos, prazo limite (SLA) e sinaliza se a solicitação está em atraso.
    """
    if cfg is None:
        cfg = get_config()
    hoje = date.today()
    atrasadas_count = 0
    visitas_processadas = []

    for v in visitas:
        v_dict = dict(v) if not isinstance(v, dict) else dict(v)
        motivo = v_dict.get('motivo') or 'Padrão'
        
        # Buscar prazo específico do motivo ou prazo padrão
        prazo_key = f"prazo_visita_{motivo}"
        prazo_str = cfg.get(prazo_key, cfg.get('prazo_visita_padrao', '15'))
        try:
            prazo_dias = int(prazo_str)
        except (ValueError, TypeError):
            prazo_dias = 15

        criado_str = str(v_dict.get('criado_em', ''))[:10]
        try:
            dt_criado = datetime.strptime(criado_str, '%Y-%m-%d').date()
            dias_decorridos = (hoje - dt_criado).days
        except Exception:
            dias_decorridos = 0

        atrasada = (v_dict.get('status') == 'Pendente') and (dias_decorridos > prazo_dias)
        dias_atraso = (dias_decorridos - prazo_dias) if atrasada else 0

        if atrasada:
            atrasadas_count += 1

        v_dict['dias_decorridos'] = dias_decorridos
        v_dict['prazo_dias'] = prazo_dias
        v_dict['atrasada'] = atrasada
        v_dict['dias_atraso'] = dias_atraso
        visitas_processadas.append(v_dict)

    return visitas_processadas, atrasadas_count


# ---------------------------------------------------------------------------
# Validação de arquivos — fotos e parecer AS
# ---------------------------------------------------------------------------

_EXTENSOES_FOTO     = {'.jpg', '.jpeg', '.png', '.webp'}
_EXTENSAO_PARECER   = '.pdf'
_TAMANHO_MAX_FOTO   = 10 * 1024 * 1024   # 10 MB
_TAMANHO_MAX_PARECER = 20 * 1024 * 1024  # 20 MB


def _extensao(filename: str) -> str:
    return os.path.splitext(filename)[1].lower()


def _validar_foto(arquivo) -> str | None:
    ext = _extensao(arquivo.filename)
    if ext not in _EXTENSOES_FOTO:
        return f"Formato não suportado: '{arquivo.filename}'. Use JPG, PNG ou WEBP."
    arquivo.seek(0, 2)
    tamanho = arquivo.tell()
    arquivo.seek(0)
    if tamanho > _TAMANHO_MAX_FOTO:
        return f"A imagem '{arquivo.filename}' excede o limite de 10 MB."
    return None


def _validar_parecer(arquivo) -> str | None:
    if _extensao(arquivo.filename) != _EXTENSAO_PARECER:
        return "O parecer deve ser um arquivo PDF."
    arquivo.seek(0, 2)
    tamanho = arquivo.tell()
    arquivo.seek(0)
    if tamanho > _TAMANHO_MAX_PARECER:
        return "O arquivo PDF excede o limite de 20 MB."
    return None


# ---------------------------------------------------------------------------
# Geração de PDF — solicitação de visita
# ---------------------------------------------------------------------------

def _logo_path(filename: str):
    """Retorna o caminho absoluto da logo se existir, senão None."""
    path = os.path.join(BASE_DIR, 'static', 'logos', filename)
    return path if os.path.exists(path) else None


def _build_pdf_story(visita, solicitante, responsavel, cfg: dict) -> list:
    if isinstance(visita, (sqlite3.Row, dict)):
        visita = dict(visita)
    if isinstance(solicitante, (sqlite3.Row, dict)):
        solicitante = dict(solicitante)
    if isinstance(responsavel, (sqlite3.Row, dict)):
        responsavel = dict(responsavel)

    styles = getSampleStyleSheet()
    verde_escuro = rl_colors.HexColor('#00542A')
    verde_vibrante = rl_colors.HexColor('#00883A')
    roxo = rl_colors.HexColor('#783DB2')
    cinza_fundo = rl_colors.HexColor('#F8F9FA')
    verde_claro = rl_colors.HexColor('#F0FDF4')
    cinza_t = rl_colors.HexColor('#4B5563')
    branco = rl_colors.white

    def par(text, size=10, bold=False, color=rl_colors.black, align='LEFT'):
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
        al = {'LEFT': TA_LEFT, 'CENTER': TA_CENTER, 'RIGHT': TA_RIGHT, 'JUSTIFY': TA_JUSTIFY}.get(align, TA_LEFT)
        st = ParagraphStyle('x', fontSize=size, fontName='Helvetica-Bold' if bold else 'Helvetica',
                            textColor=color, alignment=al, leading=size * 1.3)
        return Paragraph(str(text or ''), st)

    story = []

    # ── Timbrado / Cabeçalho de Imagem Oficial ────────────────────────────────
    img_hdr_path = os.path.join(app.root_path, 'static', 'report_assets', 'image3.png')
    if not os.path.exists(img_hdr_path):
        img_hdr_path = os.path.join(BASE_DIR, 'static', 'report_assets', 'image3.png')
    
    if os.path.exists(img_hdr_path):
        try:
            story.append(RLImage(img_hdr_path, width=17*cm, height=2.8*cm))
            story.append(Spacer(1, 0.3*cm))
        except Exception:
            pass

    # ── Título do Documento (Configurável pelo Admin) ─────────────────────────
    titulo_doc = cfg.get('visita_titulo_doc', 'SOLICITAÇÃO DE VISITA DOMICILIAR').upper()
    subtitulo_doc = cfg.get('visita_subtitulo_doc', 'Secretaria Municipal de Assistência Social - Setor do Cadastro Único / PBF')

    titulo_tbl = Table([
        [par(titulo_doc, size=12, bold=True, color=verde_escuro, align='CENTER')],
        [par(subtitulo_doc, size=9, bold=True, color=cinza_t, align='CENTER')],
    ], colWidths=[17*cm])
    story.append(titulo_tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Barra Número VD e Status Oficial ─────────────────────────────────────
    numero_vd = visita.get('numero_vd') or f"#{visita['id']}"
    status    = visita.get('status', 'Pendente')

    barra = Table([[
        par(f'Nº DA SOLICITAÇÃO: {numero_vd}', size=11, bold=True, color=branco),
        par(f'STATUS: {status.upper()}', size=11, bold=True, color=branco, align='RIGHT')
    ]], colWidths=[10*cm, 7*cm])
    barra.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), verde_escuro),
        ('TOPPADDING',  (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING',  (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(barra)
    story.append(Spacer(1, 0.4*cm))

    # ── Banner Secção 1: Dados do Beneficiário ──────────────────────────────
    hdr_sec1 = Table([[par('1. DADOS DA FAMÍLIA E LOCALIZAÇÃO', size=10, bold=True, color=branco, align='CENTER')]], colWidths=[17*cm])
    hdr_sec1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), verde_escuro),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(hdr_sec1)

    end_parts = []
    for p in [visita.get('logradouro'), visita.get('numero'), visita.get('complemento')]:
        if p: end_parts.append(p)
    endereco = ', '.join(end_parts) if end_parts else '—'

    tels = ' / '.join(filter(None, [visita.get('telefone1'), visita.get('telefone2')])) or '—'

    dados = [
        [par('CPF DO RESPONSÁVEL FAMILIAR (RF)', size=8, bold=True, color=cinza_t),
         par('NOME COMPLETO DO RESPONSÁVEL FAMILIAR', size=8, bold=True, color=cinza_t)],
        [par(visita.get('cpf_rf', '—'), size=10, bold=True),
         par(visita.get('nome_rf', '—'), size=10, bold=True)],
        [par('LOGRADOURO / ENDEREÇO', size=8, bold=True, color=cinza_t),
         par('BAIRRO / ZONA', size=8, bold=True, color=cinza_t)],
        [par(endereco, size=9),
         par(f"{visita.get('bairro', '—')} ({visita.get('zona', 'Urbana')})", size=9)],
        [par('PONTO DE REFERÊNCIA', size=8, bold=True, color=cinza_t),
         par('TELEFONES DE CONTATO', size=8, bold=True, color=cinza_t)],
        [par(visita.get('referencia') or '—', size=9),
         par(tels, size=9)],
    ]
    dt = Table(dados, colWidths=[8.5*cm, 8.5*cm])
    dt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), verde_claro),
        ('BACKGROUND', (0, 2), (-1, 2), verde_claro),
        ('BACKGROUND', (0, 4), (-1, 4), verde_claro),
        ('BOX',    (0, 0), (-1, -1), 0.5, verde_escuro),
        ('GRID',   (0, 0), (-1, -1), 0.3, rl_colors.HexColor('#D1D5DB')),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
    ]))
    story.append(dt)
    story.append(Spacer(1, 0.4*cm))

    # ── Banner Secção 2: Motivo e Dados Administrativos ─────────────────────
    hdr_sec2 = Table([[par('2. DETALHES DA SOLICITAÇÃO E MOTIVO', size=10, bold=True, color=branco, align='CENTER')]], colWidths=[17*cm])
    hdr_sec2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), verde_escuro),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(hdr_sec2)

    criado_raw = visita.get('criado_em', '')
    criado_fmt = ''
    if criado_raw:
        try:
            criado_fmt = datetime.fromisoformat(criado_raw[:19]).strftime('%d/%m/%Y %H:%M')
        except Exception:
            criado_fmt = criado_raw[:10]

    sol_nome  = (dict(solicitante).get('nome')) if solicitante else '—'
    resp_nome = (dict(responsavel).get('nome')) if responsavel else 'Não atribuído'

    adm_dados = [
        [par('MOTIVO PRINCIPAL DA VISITA', size=8, bold=True, color=cinza_t),
         par('DATA DA SOLICITAÇÃO', size=8, bold=True, color=cinza_t)],
        [par(visita.get('motivo', '—'), size=10, bold=True, color=verde_escuro),
         par(criado_fmt, size=9)],
        [par('SOLICITANTE DO REGISTRO', size=8, bold=True, color=cinza_t),
         par('ENTREVISTADOR / AS RESPONSÁVEL', size=8, bold=True, color=cinza_t)],
        [par(sol_nome, size=9),
         par(resp_nome, size=9, bold=True)],
    ]
    at = Table(adm_dados, colWidths=[9.5*cm, 7.5*cm])
    at.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), verde_claro),
        ('BACKGROUND', (0, 2), (-1, 2), verde_claro),
        ('BOX',   (0, 0), (-1, -1), 0.5, verde_escuro),
        ('GRID',  (0, 0), (-1, -1), 0.3, rl_colors.HexColor('#D1D5DB')),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
    ]))
    story.append(at)
    story.append(Spacer(1, 0.4*cm))

    # ── Observações da Visita ────────────────────────────────────────────────
    obs = visita.get('observacoes')
    if obs:
        story.append(par('OBSERVAÇÕES ADICIONAIS:', size=9, bold=True, color=verde_escuro))
        story.append(par(obs, size=9))
        story.append(Spacer(1, 0.3*cm))

    # ── Parecer Técnico Assistencial (Serviço Social) ────────────────────────
    parecer_txt = visita.get('parecer_tecnico_txt')
    if parecer_txt:
        hdr_parecer = Table([[par('PARECER TÉCNICO ASSISTENCIAL / SERVIÇO SOCIAL', size=9, bold=True, color=branco, align='CENTER')]], colWidths=[17*cm])
        hdr_parecer.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), verde_escuro),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(hdr_parecer)

        box_parecer = Table([[par(parecer_txt, size=8.5, color=cinza_t, align='JUSTIFY')]], colWidths=[17*cm])
        box_parecer.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), cinza_fundo),
            ('BOX', (0, 0), (-1, -1), 0.5, verde_escuro),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(box_parecer)
        story.append(Spacer(1, 0.4*cm))

    # ── Texto de Orientação Institucional (Configurável no Admin) ────────────
    orientacao_txt = cfg.get('visita_orientacao_texto')
    if orientacao_txt:
        hdr_orient = Table([[par('ORIENTAÇÕES INSTITUCIONAIS PARA A VISITA', size=9, bold=True, color=branco, align='CENTER')]], colWidths=[17*cm])
        hdr_orient.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), roxo),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(hdr_orient)

        box_orient = Table([[par(orientacao_txt, size=8.5, color=cinza_t, align='JUSTIFY')]], colWidths=[17*cm])
        box_orient.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), cinza_fundo),
            ('BOX', (0, 0), (-1, -1), 0.5, roxo),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(box_orient)
        story.append(Spacer(1, 0.4*cm))

    # ── Assinaturas Oficiais (Configuráveis no Admin) ────────────────────────
    story.append(Spacer(1, 0.8*cm))
    lbl_ass1 = cfg.get('visita_assinatura_1', 'Assinatura do Entrevistador / Assistente Social')
    lbl_ass2 = cfg.get('visita_assinatura_2', 'Assinatura do Responsável Familiar (RF)')

    sig = Table([
        [par('', size=14), par('', size=14), par('', size=14)],
        [par(lbl_ass1, size=8.5, bold=True, align='CENTER'),
         par('', size=8.5),
         par(lbl_ass2, size=8.5, bold=True, align='CENTER')],
        [par(resp_nome, size=8, color=cinza_t, align='CENTER'),
         par('', size=8),
         par(visita.get('nome_rf', ''), size=8, color=cinza_t, align='CENTER')],
    ], colWidths=[7.5*cm, 2*cm, 7.5*cm])
    sig.setStyle(TableStyle([
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'BOTTOM'),
        ('LINEBELOW',     (0, 0), (0, 0),   1.0, verde_escuro),
        ('LINEBELOW',     (2, 0), (2, 0),   1.0, verde_escuro),
        ('TOPPADDING',    (0, 0), (-1, 0),  16),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  2),
    ]))
    story.append(sig)
    story.append(Spacer(1, 0.6*cm))

    # ── Rodapé Oficial ───────────────────────────────────────────────────────
    rodape_txt = cfg.get('visita_rodape_txt', cfg.get('rodape', ''))
    municipio  = cfg.get('municipio', 'Tomé-Açu/PA')
    conteudo_rodape = f"{cfg.get('setor_nome', 'Setor do Cadastro Único / PBF')}  |  {rodape_txt}  |  {municipio}"
    
    rodape_tbl = Table([[par(conteudo_rodape, size=7.5, color=branco, align='CENTER')]], colWidths=[17*cm])
    rodape_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), verde_escuro),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(rodape_tbl)

    return story


def gerar_pdf_visita(visita_id: int):
    """Gera o PDF da solicitação e retorna (bytes, numero_vd) ou (None, None)."""
    conn = get_db()
    visita = _fetchone(conn, "SELECT * FROM solicitacoes_visita WHERE id=?", (visita_id,))
    if not visita:
        conn.close()
        return None, None
    solicitante = _fetchone(conn, "SELECT nome FROM usuarios WHERE id=?",
                            (visita['solicitante_id'],))
    responsavel = None
    if visita['responsavel_id']:
        responsavel = _fetchone(conn, "SELECT nome FROM usuarios WHERE id=?",
                                (visita['responsavel_id'],))
    cfg = get_config()
    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    story = _build_pdf_story(visita, solicitante, responsavel, cfg)
    doc.build(story)
    numero_vd = (dict(visita).get('numero_vd')) or f"visita-{visita_id}"
    return buf.getvalue(), numero_vd


def _gerar_numero_vd(conn, ano: int) -> str:
    """
    Incrementa atomicamente o contador de VD para o ano e retorna
    o número no formato 'VD-AAAA-NNNNNN'.
    Lança ValueError('limite_anual') se o contador atingir 999999.
    """
    if _is_pg():
        row = _fetchone(conn,
            "SELECT ultimo_numero FROM visita_contadores WHERE ano = %s FOR UPDATE",
            (ano,)
        )
        if row is None:
            _exec(conn,
                "INSERT INTO visita_contadores (ano, ultimo_numero) VALUES (%s, 1)",
                (ano,)
            )
            proximo = 1
        else:
            proximo = row['ultimo_numero'] + 1
            if proximo > 999999:
                raise ValueError("limite_anual")
            _exec(conn,
                "UPDATE visita_contadores SET ultimo_numero = %s WHERE ano = %s",
                (proximo, ano)
            )
    else:
        row = _fetchone(conn,
            "SELECT ultimo_numero FROM visita_contadores WHERE ano = ?",
            (ano,)
        )
        if row is None:
            _exec(conn,
                "INSERT INTO visita_contadores (ano, ultimo_numero) VALUES (?, 1)",
                (ano,)
            )
            proximo = 1
        else:
            proximo = row['ultimo_numero'] + 1
            if proximo > 999999:
                raise ValueError("limite_anual")
            _exec(conn,
                "UPDATE visita_contadores SET ultimo_numero = ? WHERE ano = ?",
                (proximo, ano)
            )
    return f"VD-{ano}-{proximo:06d}"


def init_db():
    conn = get_db()
    if _is_pg():
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            login TEXT UNIQUE NOT NULL,
            senha TEXT NOT NULL,
            perfil TEXT NOT NULL DEFAULT 'entrevistador',
            acesso_sibec INTEGER NOT NULL DEFAULT 0,
            trocar_senha INTEGER NOT NULL DEFAULT 0,
            email TEXT,
            telefone TEXT
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS atendimentos (
            id SERIAL PRIMARY KEY,
            data TEXT NOT NULL,
            cpf TEXT NOT NULL,
            nome_rf TEXT NOT NULL,
            origem TEXT NOT NULL,
            tipos TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            criado_em TEXT NOT NULL
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            usuario_id INTEGER,
            usuario_nome TEXT,
            acao TEXT NOT NULL,
            detalhe TEXT,
            ip TEXT,
            criado_em TEXT NOT NULL
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS config_relatorio (
            chave TEXT PRIMARY KEY,
            valor TEXT NOT NULL
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS solicitacoes_visita (
            id                  SERIAL PRIMARY KEY,
            cpf_rf              TEXT NOT NULL,
            nome_rf             TEXT NOT NULL,
            logradouro          TEXT NOT NULL DEFAULT '',
            numero              TEXT NOT NULL DEFAULT '',
            complemento         TEXT,
            bairro              TEXT NOT NULL DEFAULT '',
            referencia          TEXT,
            zona                TEXT NOT NULL DEFAULT 'Urbana',
            motivo              TEXT NOT NULL,
            data_realizada      TEXT,
            status              TEXT NOT NULL DEFAULT 'Pendente',
            solicitante_id      INTEGER NOT NULL REFERENCES usuarios(id),
            responsavel_id      INTEGER REFERENCES usuarios(id),
            observacoes         TEXT,
            motivo_cancelamento TEXT,
            anexo_url           TEXT,
            anexo_nome          TEXT,
            atendimento_id      INTEGER,
            criado_em           TEXT NOT NULL,
            atualizado_em       TEXT NOT NULL
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS visita_contadores (
            ano             INTEGER PRIMARY KEY,
            ultimo_numero   INTEGER NOT NULL DEFAULT 0
        )''')
        cur.execute('''CREATE TABLE IF NOT EXISTS visita_fotos (
            id              SERIAL PRIMARY KEY,
            solicitacao_id  INTEGER NOT NULL REFERENCES solicitacoes_visita(id) ON DELETE CASCADE,
            url             TEXT NOT NULL,
            nome_arquivo    TEXT NOT NULL,
            criado_em       TEXT NOT NULL
        )''')
        # Migrações seguras para PostgreSQL — cada uma em savepoint individual
        migracoes = [
            "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS unidade TEXT DEFAULT 'Tomé-Açu (Sede)'",
            "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS email TEXT",
            "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS telefone TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS atendimento_id INTEGER",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS anexo_url TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS anexo_nome TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS logradouro TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS numero TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS complemento TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS bairro TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS referencia TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS zona TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS motivo_especificado TEXT",
            "ALTER TABLE solicitacoes_visita ALTER COLUMN endereco DROP NOT NULL",
            "ALTER TABLE solicitacoes_visita ALTER COLUMN motivo DROP NOT NULL",
            "ALTER TABLE solicitacoes_visita ALTER COLUMN data_prevista DROP NOT NULL",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS numero_vd TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS parecer_as_url TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS parecer_as_nome TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS telefone1 TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS telefone2 TEXT",
            "ALTER TABLE atendimentos ADD COLUMN IF NOT EXISTS orgao_encaminhador TEXT",
            "ALTER TABLE atendimentos ADD COLUMN IF NOT EXISTS orgao_outro TEXT",
            "ALTER TABLE atendimentos ADD COLUMN IF NOT EXISTS numero_oficio TEXT",
            "ALTER TABLE atendimentos ADD COLUMN IF NOT EXISTS data_encaminhamento TEXT",
            "ALTER TABLE atendimentos ADD COLUMN IF NOT EXISTS servidor_encaminhador TEXT",
            "ALTER TABLE atendimentos ADD COLUMN IF NOT EXISTS motivo_encaminhamento TEXT",
            "ALTER TABLE atendimentos ADD COLUMN IF NOT EXISTS obs_encaminhamento TEXT",
            "ALTER TABLE atendimentos ADD COLUMN IF NOT EXISTS situacao_encaminhamento TEXT DEFAULT 'Atendido'",
            "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS tentativas_login INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE solicitacoes_visita ADD COLUMN IF NOT EXISTS parecer_tecnico_txt TEXT",
        ]
        for i, col_sql in enumerate(migracoes):
            sp = f"sp_mig_{i}"
            try:
                cur.execute(f"SAVEPOINT {sp}")
                cur.execute(col_sql)
                cur.execute(f"RELEASE SAVEPOINT {sp}")
            except Exception:
                cur.execute(f"ROLLBACK TO SAVEPOINT {sp}")
        # Insere configurações padrão do relatório se não existirem
        defaults = _config_defaults()
        for chave, valor in defaults.items():
            cur.execute(
                "INSERT INTO config_relatorio (chave, valor) VALUES (%s, %s) ON CONFLICT (chave) DO NOTHING",
                (chave, valor)
            )
        row = _fetchone(conn, "SELECT COUNT(*) as n FROM usuarios WHERE login='admin'")
        if row['n'] == 0:
            cur.execute(
                "INSERT INTO usuarios (nome,login,senha,perfil,acesso_sibec,trocar_senha) VALUES (%s,%s,%s,%s,%s,%s)",
                ('Administrador', 'admin', generate_password_hash('admin123'), 'admin', 1, 1)
            )
        conn.commit()
        conn.close()
    else:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            login TEXT UNIQUE NOT NULL,
            senha TEXT NOT NULL,
            perfil TEXT NOT NULL DEFAULT 'entrevistador',
            acesso_sibec INTEGER NOT NULL DEFAULT 0,
            trocar_senha INTEGER NOT NULL DEFAULT 0,
            tentativas_login INTEGER NOT NULL DEFAULT 0
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS atendimentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            cpf TEXT NOT NULL,
            nome_rf TEXT NOT NULL,
            origem TEXT NOT NULL,
            tipos TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            criado_em TEXT NOT NULL,
            orgao_encaminhador TEXT,
            orgao_outro TEXT,
            numero_oficio TEXT,
            data_encaminhamento TEXT,
            servidor_encaminhador TEXT,
            motivo_encaminhamento TEXT,
            obs_encaminhamento TEXT,
            situacao_encaminhamento TEXT DEFAULT 'Atendido',
            FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER,
            usuario_nome TEXT,
            acao TEXT NOT NULL,
            detalhe TEXT,
            ip TEXT,
            criado_em TEXT NOT NULL
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS config_relatorio (
            chave TEXT PRIMARY KEY,
            valor TEXT NOT NULL
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS solicitacoes_visita (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            cpf_rf              TEXT NOT NULL,
            nome_rf             TEXT NOT NULL,
            logradouro          TEXT NOT NULL DEFAULT '',
            numero              TEXT NOT NULL DEFAULT '',
            complemento         TEXT,
            bairro              TEXT NOT NULL DEFAULT '',
            referencia          TEXT,
            zona                TEXT NOT NULL DEFAULT 'Urbana',
            motivo              TEXT NOT NULL,
            data_realizada      TEXT,
            status              TEXT NOT NULL DEFAULT 'Pendente',
            solicitante_id      INTEGER NOT NULL REFERENCES usuarios(id),
            responsavel_id      INTEGER REFERENCES usuarios(id),
            observacoes         TEXT,
            motivo_cancelamento TEXT,
            anexo_url           TEXT,
            anexo_nome          TEXT,
            atendimento_id      INTEGER,
            criado_em           TEXT NOT NULL,
            atualizado_em       TEXT NOT NULL,
            parecer_tecnico_txt TEXT
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS visita_contadores (
            ano             INTEGER PRIMARY KEY,
            ultimo_numero   INTEGER NOT NULL DEFAULT 0
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS visita_fotos (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            solicitacao_id  INTEGER NOT NULL REFERENCES solicitacoes_visita(id) ON DELETE CASCADE,
            url             TEXT NOT NULL,
            nome_arquivo    TEXT NOT NULL,
            criado_em       TEXT NOT NULL
        )''')
        for col_sql in [
            "ALTER TABLE usuarios ADD COLUMN unidade TEXT DEFAULT 'Tomé-Açu (Sede)'",
            "ALTER TABLE usuarios ADD COLUMN trocar_senha INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE usuarios ADD COLUMN email TEXT",
            "ALTER TABLE usuarios ADD COLUMN telefone TEXT",
            "ALTER TABLE usuarios ADD COLUMN tentativas_login INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE solicitacoes_visita ADD COLUMN numero_vd TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN parecer_as_url TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN parecer_as_nome TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN parecer_tecnico_txt TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN telefone1 TEXT",
            "ALTER TABLE solicitacoes_visita ADD COLUMN telefone2 TEXT",
            "ALTER TABLE atendimentos ADD COLUMN orgao_encaminhador TEXT",
            "ALTER TABLE atendimentos ADD COLUMN orgao_outro TEXT",
            "ALTER TABLE atendimentos ADD COLUMN numero_oficio TEXT",
            "ALTER TABLE atendimentos ADD COLUMN data_encaminhamento TEXT",
            "ALTER TABLE atendimentos ADD COLUMN servidor_encaminhador TEXT",
            "ALTER TABLE atendimentos ADD COLUMN motivo_encaminhamento TEXT",
            "ALTER TABLE atendimentos ADD COLUMN obs_encaminhamento TEXT",
            "ALTER TABLE atendimentos ADD COLUMN situacao_encaminhamento TEXT DEFAULT 'Atendido'",
        ]:
            try:
                c.execute(col_sql)
            except Exception:
                pass
        # Insere configurações padrão se não existirem
        defaults = _config_defaults()
        for chave, valor in defaults.items():
            try:
                c.execute("INSERT INTO config_relatorio (chave, valor) VALUES (?, ?)", (chave, valor))
            except Exception:
                pass
        row = _fetchone(conn, "SELECT COUNT(*) FROM usuarios WHERE login='admin'")
        count = row[0] if row else 0
        if count == 0:
            c.execute(
                "INSERT INTO usuarios (nome,login,senha,perfil,acesso_sibec,trocar_senha) VALUES (?,?,?,?,?,?)",
                ('Administrador', 'admin', generate_password_hash('admin123'), 'admin', 1, 1)
            )
        conn.commit()
        conn.close()


def audit(acao: str, detalhe: str = ''):
    """Registra uma ação no log de auditoria."""
    conn = get_db()
    _exec(conn,
        "INSERT INTO audit_log (usuario_id,usuario_nome,acao,detalhe,ip,criado_em) VALUES (?,?,?,?,?,?)",
        (
            session.get('usuario_id'),
            session.get('usuario_nome', '?'),
            acao, detalhe,
            request.remote_addr,
            datetime.now(_TZ_BELEM).isoformat(),
        ),
    )
    conn.commit()
    conn.close()

# ---------------------------------------------------------------------------
# Autenticação
# ---------------------------------------------------------------------------

def _requer_login():
    return 'usuario_id' not in session


@app.route('/')
def index():
    if _requer_login():
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))


@app.route('/tutorial')
def tutorial():
    return redirect(url_for('dashboard'))


@app.context_processor
def inject_notificacoes_visitas():
    if 'usuario_id' not in session:
        return {'notif_visitas_atrasadas': [], 'notif_total_atrasadas': 0}
    try:
        conn = get_db()
        uid = session['usuario_id']
        perfil = session.get('perfil')
        if perfil == 'admin':
            visitas = _fetchall(conn, "SELECT * FROM solicitacoes_visita WHERE status='Pendente' ORDER BY criado_em ASC")
        else:
            visitas = _fetchall(conn, "SELECT * FROM solicitacoes_visita WHERE status='Pendente' AND (solicitante_id=? OR responsavel_id=?) ORDER BY criado_em ASC", (uid, uid))
        cfg = get_config()
        conn.close()
        visitas_proc, total_atrasadas = _processar_sla_visitas(visitas, cfg)
        visitas_atrasadas = [v for v in visitas_proc if v.get('atrasada')]
        return {
            'notif_visitas_atrasadas': visitas_atrasadas,
            'notif_total_atrasadas': len(visitas_atrasadas)
        }
    except Exception:
        return {'notif_visitas_atrasadas': [], 'notif_total_atrasadas': 0}


import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


def _enviar_email_smtp(destino, assunto, html_corpo):
    """Envia um e-mail HTML via SMTP usando as configurações cadastradas no sistema."""
    cfg = get_config()
    smtp_host = cfg.get('smtp_host', 'smtp.gmail.com').strip()
    try:
        smtp_port = int(cfg.get('smtp_port', 587))
    except (ValueError, TypeError):
        smtp_port = 587
    smtp_user = cfg.get('smtp_user', cfg.get('email_setor', 'setascadastrounico@gmail.com')).strip()
    smtp_pass = cfg.get('smtp_pass', '').strip()

    if not smtp_user or not smtp_pass:
        return False, "Configurações de e-mail (SMTP) não preenchidas no sistema."

    msg = MIMEMultipart('alternative')
    msg['Subject'] = assunto
    msg['From'] = f"Sistema CadÚnico <{smtp_user}>"
    msg['To'] = destino
    msg.attach(MIMEText(html_corpo, 'html'))

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, destino, msg.as_string())
        return True, "E-mail enviado com sucesso."
    except Exception as e:
        return False, f"Falha no envio via SMTP: {str(e)}"


def _notificar_admin_senha_incorreta(usuario, tentativas):
    """Notifica o e-mail do administrador quando um usuário erra a senha 5 ou mais vezes."""
    cfg = get_config()
    email_admin = cfg.get('email_admin_notificacao') or cfg.get('email_setor') or 'setascadastrounico@gmail.com'
    u_dict = dict(usuario)
    nome_user = u_dict.get('nome', 'Usuário')
    login_user = u_dict.get('login', 'N/A')
    horario = datetime.now(_TZ_BELEM).strftime('%d/%m/%Y às %H:%M:%S')

    assunto = f"⚠️ [Alerta CadÚnico] 5 Tentativas Incorretas de Senha - {nome_user}"
    html = f"""
    <div style="font-family: Arial, sans-serif; max-width:600px; margin:0 auto; padding:20px; border:1px solid #E2E8F0; border-radius:12px; background:#FFFFFF;">
      <div style="background:#00542A; color:#FFFFFF; padding:16px; border-radius:8px 8px 0 0; text-align:center;">
        <h2 style="margin:0; font-size:18px; color:#FFFFFF;">⚠️ Alerta de Segurança — Tentativas de Senha</h2>
        <p style="margin:4px 0 0 0; font-size:13px; opacity:0.9;">Sistema de Cadastro Único</p>
      </div>
      <div style="padding:20px; color:#1E293B; line-height:1.6;">
        <p>Olá, <strong>Administrador</strong>,</p>
        <p>O usuário <strong>{nome_user}</strong> (Login: <code>{login_user}</code>) errou a senha de acesso <strong>{tentativas} vezes seguidas</strong>.</p>
        <div style="background:#FFFBEB; border-left:4px solid #FAB614; padding:12px 16px; margin:16px 0; border-radius:4px;">
          <strong>Data/Horário:</strong> {horario}<br>
          <strong>Total de Falhas:</strong> {tentativas} tentativas incorretas
        </div>
        <p>Acesse o painel administrativo do sistema para efetuar a redefinição da senha do usuário ou orientar o seu acesso.</p>
      </div>
      <div style="border-top:1px solid #E2E8F0; padding-top:12px; text-align:center; font-size:12px; color:#64748B;">
        Sistema Oficial do Cadastro Único · Notificação Automática
      </div>
    </div>
    """
    
    sucesso, msg = _enviar_email_smtp(email_admin, assunto, html)
    audit('ALERTA_SENHA_5X', f"usuario={login_user} email_admin={email_admin} env_status={sucesso}")
    return sucesso


@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        login_ = request.form.get('login', '').strip()
        senha = request.form.get('senha', '').strip()
        conn = get_db()
        u = _fetchone(conn, "SELECT * FROM usuarios WHERE login=?", (login_,))
        if u and check_password_hash(u['senha'], senha):
            _exec(conn, "UPDATE usuarios SET tentativas_login=0 WHERE id=?", (u['id'],))
            conn.commit()
            conn.close()
            session['usuario_id'] = u['id']
            session['usuario_nome'] = u['nome']
            session['perfil'] = u['perfil']
            session['acesso_sibec'] = bool(u['acesso_sibec'])
            session['trocar_senha'] = bool(u['trocar_senha'])
            audit('LOGIN', f"login={login_}")
            if u['trocar_senha']:
                return redirect(url_for('trocar_senha_obrigatorio'))
            return redirect(url_for('dashboard'))
        else:
            if u:
                u_dict = dict(u)
                novas_tentativas = (u_dict.get('tentativas_login') or 0) + 1
                _exec(conn, "UPDATE usuarios SET tentativas_login=? WHERE id=?", (novas_tentativas, u['id']))
                conn.commit()
                conn.close()
                audit('LOGIN_FALHA', f"login={login_} tentativas={novas_tentativas}")
                if novas_tentativas >= 5:
                    _notificar_admin_senha_incorreta(u, novas_tentativas)
                    erro = f"Login ou senha incorretos. Você errou a senha {novas_tentativas} vezes. O Administrador foi notificado no e-mail para auxiliar na troca de senha."
                else:
                    erro = 'Login ou senha incorretos.'
            else:
                conn.close()
                erro = 'Login ou senha incorretos.'
    return render_template('login.html', erro=erro)


@app.route('/logout')
def logout():
    audit('LOGOUT')
    session.clear()
    return redirect(url_for('login'))


@app.route('/trocar-senha', methods=['GET', 'POST'])
def trocar_senha_obrigatorio():
    if _requer_login():
        return redirect(url_for('login'))
    erros = []
    if request.method == 'POST':
        nova = request.form.get('nova_senha', '').strip()
        conf = request.form.get('confirmar_senha', '').strip()
        if len(nova) < 6:
            erros.append('A senha deve ter ao menos 6 caracteres.')
        if nova != conf:
            erros.append('As senhas não coincidem.')
        if not erros:
            conn = get_db()
            _exec(conn,
                "UPDATE usuarios SET senha=?, trocar_senha=0 WHERE id=?",
                (generate_password_hash(nova), session['usuario_id'])
            )
            conn.commit()
            conn.close()
            session['trocar_senha'] = False
            audit('TROCA_SENHA', 'senha alterada por obrigação no primeiro acesso')
            flash('Senha alterada com sucesso!', 'ok')
            return redirect(url_for('dashboard'))
    return render_template('trocar_senha.html', erros=erros)


# ---------------------------------------------------------------------------
# Recuperação de senha (token + WhatsApp / e-mail)
# ---------------------------------------------------------------------------

import secrets

# Armazena tokens em memória: {token: {'usuario_id': int, 'expira': datetime}}
_tokens_recuperacao: dict = {}

TOKEN_VALIDADE_MIN = 30  # minutos


def _gerar_token_recuperacao(usuario_id: int) -> str:
    token = secrets.token_urlsafe(32)
    _tokens_recuperacao[token] = {
        'usuario_id': usuario_id,
        'expira': datetime.now() + timedelta(minutes=TOKEN_VALIDADE_MIN),
    }
    return token


def _validar_token(token: str):
    """Retorna usuario_id se token válido, None caso contrário."""
    dado = _tokens_recuperacao.get(token)
    if not dado:
        return None
    if datetime.now() > dado['expira']:
        _tokens_recuperacao.pop(token, None)
        return None
    return dado['usuario_id']


# ---------------------------------------------------------------------------
# Meu Perfil — dados pessoais + estatísticas
# ---------------------------------------------------------------------------

@app.route('/meu-perfil', methods=['GET', 'POST'])
def meu_perfil():
    if _requer_login():
        return redirect(url_for('login'))

    uid = session['usuario_id']
    conn = get_db()
    u = _fetchone(conn, "SELECT * FROM usuarios WHERE id=?", (uid,))
    erros = []
    secao = request.args.get('secao', 'dados')  # 'dados' ou 'senha'

    if request.method == 'POST':
        acao = request.form.get('acao', 'dados')

        if acao == 'dados':
            nome = request.form.get('nome', '').strip()
            email = request.form.get('email', '').strip() or None
            telefone = request.form.get('telefone', '').strip() or None
            if not nome:
                erros.append('Nome obrigatório.')
            if not erros:
                _exec(conn,
                    "UPDATE usuarios SET nome=?, email=?, telefone=? WHERE id=?",
                    (nome, email, telefone, uid)
                )
                conn.commit()
                session['usuario_nome'] = nome
                audit('PERFIL_DADOS', f"nome={nome}")
                flash('Dados atualizados com sucesso!', 'ok')
                conn.close()
                return redirect(url_for('meu_perfil', secao='dados'))

        elif acao == 'senha':
            senha_atual = request.form.get('senha_atual', '').strip()
            nova = request.form.get('nova_senha', '').strip()
            conf = request.form.get('confirmar_senha', '').strip()
            if not check_password_hash(u['senha'], senha_atual):
                erros.append('Senha atual incorreta.')
            if len(nova) < 6:
                erros.append('A nova senha deve ter ao menos 6 caracteres.')
            if nova != conf:
                erros.append('As senhas não coincidem.')
            if not erros:
                _exec(conn,
                    "UPDATE usuarios SET senha=?, trocar_senha=0 WHERE id=?",
                    (generate_password_hash(nova), uid)
                )
                conn.commit()
                conn.close()
                session['trocar_senha'] = False
                audit('PERFIL_SENHA', 'senha alterada pelo próprio usuário')
                flash('Senha alterada com sucesso!', 'ok')
                return redirect(url_for('meu_perfil', secao='senha'))

        # Recarrega usuário após possível commit parcial
        u = _fetchone(conn, "SELECT * FROM usuarios WHERE id=?", (uid,))
        secao = acao

    # ── Estatísticas ──────────────────────────────────────────────────────
    hoje = date.today()
    mes_atual = hoje.strftime('%Y-%m')
    ano_atual = str(hoje.year)

    total_hoje = _fetchone(conn,
        "SELECT COUNT(*) as n FROM atendimentos WHERE usuario_id=? AND data=?",
        (uid, hoje.isoformat())
    )['n']

    total_mes = _fetchone(conn,
        "SELECT COUNT(*) as n FROM atendimentos WHERE usuario_id=? AND data LIKE ?",
        (uid, mes_atual + '%')
    )['n']

    total_ano = _fetchone(conn,
        "SELECT COUNT(*) as n FROM atendimentos WHERE usuario_id=? AND data LIKE ?",
        (uid, ano_atual + '%')
    )['n']

    total_geral = _fetchone(conn,
        "SELECT COUNT(*) as n FROM atendimentos WHERE usuario_id=?", (uid,)
    )['n']

    # Atendimentos por mês nos últimos 12 meses
    if _is_pg():
        meses_rows = _fetchall(conn, """
            SELECT substring(data from 1 for 7) as mes, COUNT(*) as total
            FROM atendimentos WHERE usuario_id=%s
            AND data >= to_char(NOW() - INTERVAL '11 months', 'YYYY-MM') || '-01'
            GROUP BY mes ORDER BY mes
        """, (uid,))
    else:
        meses_rows = _fetchall(conn, """
            SELECT substr(data,1,7) as mes, COUNT(*) as total
            FROM atendimentos WHERE usuario_id=?
            AND data >= date('now','-11 months','start of month')
            GROUP BY mes ORDER BY mes
        """, (uid,))

    # Tipo mais frequente
    tipo_top = _fetchone(conn,
        f"SELECT tipos, COUNT(*) as n FROM atendimentos WHERE usuario_id={PH} AND data LIKE {PH} GROUP BY tipos ORDER BY n DESC LIMIT 1",
        (uid, mes_atual + '%'))

    # Dia com mais atendimentos no mês atual
    dia_top = _fetchone(conn,
        f"SELECT data, COUNT(*) as n FROM atendimentos WHERE usuario_id={PH} AND data LIKE {PH} GROUP BY data ORDER BY n DESC LIMIT 1",
        (uid, mes_atual + '%'))

    # Média diária no mês (dias com ao menos 1 atendimento)
    dias_trabalhados = _fetchone(conn,
        f"SELECT COUNT(DISTINCT data) as n FROM atendimentos WHERE usuario_id={PH} AND data LIKE {PH}",
        (uid, mes_atual + '%'))['n']

    media_dia = round(total_mes / dias_trabalhados, 1) if dias_trabalhados else 0

    # Histórico mensal completo (para tabela)
    if _is_pg():
        historico = _fetchall(conn, """
            SELECT substring(data from 1 for 7) as mes, COUNT(*) as total
            FROM atendimentos WHERE usuario_id=%s
            GROUP BY mes ORDER BY mes DESC LIMIT 24
        """, (uid,))
    else:
        historico = _fetchall(conn, """
            SELECT substr(data,1,7) as mes, COUNT(*) as total
            FROM atendimentos WHERE usuario_id=?
            GROUP BY mes ORDER BY mes DESC LIMIT 24
        """, (uid,))

    conn.close()

    # Nomes dos meses para o gráfico
    MESES_ABREV = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
    grafico_labels = []
    grafico_valores = []
    for row in meses_rows:
        try:
            dt = datetime.strptime(row['mes'] + '-01', '%Y-%m-%d')
            grafico_labels.append(MESES_ABREV[dt.month - 1] + '/' + str(dt.year)[2:])
        except Exception:
            grafico_labels.append(row['mes'])
        grafico_valores.append(row['total'])

    return render_template(
        'meu_perfil.html',
        u=u, erros=erros, secao=secao,
        total_hoje=total_hoje,
        total_mes=total_mes,
        total_ano=total_ano,
        total_geral=total_geral,
        media_dia=media_dia,
        dias_trabalhados=dias_trabalhados,
        tipo_top=tipo_top,
        dia_top=dia_top,
        historico=historico,
        grafico_labels=grafico_labels,
        grafico_valores=grafico_valores,
        mes_nome_atual=nome_mes(mes_atual), 
        ano_atual=ano_atual,
    )


@app.route('/recuperar-senha', methods=['GET', 'POST'])
def recuperar_senha():
    """Tela onde o usuário informa login para receber o link de recuperação."""
    msg = None
    erro = None
    canal = None          # 'whatsapp' ou 'email'
    link = None           # link gerado (mostrado na tela para uso imediato)
    telefone_mascarado = None
    email_mascarado = None

    if request.method == 'POST':
        login_ = request.form.get('login', '').strip()
        canal = request.form.get('canal', 'whatsapp')
        conn = get_db()
        u = _fetchone(conn, "SELECT * FROM usuarios WHERE login=?", (login_,))
        conn.close()

        if not u:
            erro = 'Usuário não encontrado.'
        else:
            token = _gerar_token_recuperacao(u['id'])
            link = url_for('redefinir_senha', token=token, _external=True)

            if canal == 'whatsapp':
                tel = u['telefone'] if u['telefone'] else None
                if not tel:
                    erro = ('Este usuário não tem telefone cadastrado. '
                            'Peça ao administrador para cadastrar ou use outro canal.')
                else:
                    # Mascara o número: ex. (91) 9****-1234
                    tel_digits = ''.join(c for c in tel if c.isdigit())
                    if len(tel_digits) >= 4:
                        telefone_mascarado = tel_digits[:2] + ' 9' + '*' * (len(tel_digits) - 6) + tel_digits[-4:]
                    else:
                        telefone_mascarado = '****'
                    numero_limpo = ''.join(c for c in tel if c.isdigit())
                    texto_wpp = (
                        f"Olá {u['nome']}! Aqui está seu link para redefinir a senha "
                        f"do sistema CadÚnico/PBF (válido por {TOKEN_VALIDADE_MIN} min):\n{link}"
                    )
                    import urllib.parse
                    wpp_url = f"https://wa.me/55{numero_limpo}?text={urllib.parse.quote(texto_wpp)}"
                    msg = 'whatsapp'
                    # Passamos o link do WhatsApp para abrir direto
                    canal = wpp_url

            elif canal == 'email':
                eml = u['email'] if u['email'] else None
                if not eml:
                    erro = ('Este usuário não tem e-mail cadastrado. '
                            'Peça ao administrador para cadastrar ou use outro canal.')
                else:
                    partes = eml.split('@')
                    nome_eml = partes[0][:2] + '*' * (len(partes[0]) - 2)
                    email_mascarado = nome_eml + '@' + (partes[1] if len(partes) > 1 else '***')
                    msg = 'email'

            if not erro:
                audit('RECUPERACAO_SENHA', f"login={login_} canal={('whatsapp' if 'wa.me' in str(canal) else canal)}")

    return render_template(
        'recuperar_senha.html',
        msg=msg, erro=erro, canal=canal, link=link,
        telefone_mascarado=telefone_mascarado,
        email_mascarado=email_mascarado,
        TOKEN_VALIDADE_MIN=TOKEN_VALIDADE_MIN,
    )


@app.route('/redefinir-senha/<token>', methods=['GET', 'POST'])
def redefinir_senha(token):
    """Tela de redefinição de senha via token."""
    uid = _validar_token(token)
    if not uid:
        flash('Link inválido ou expirado. Solicite um novo.', 'aviso')
        return redirect(url_for('recuperar_senha'))

    erros = []
    if request.method == 'POST':
        nova = request.form.get('nova_senha', '').strip()
        conf = request.form.get('confirmar_senha', '').strip()
        if len(nova) < 6:
            erros.append('A senha deve ter ao menos 6 caracteres.')
        if nova != conf:
            erros.append('As senhas não coincidem.')
        if not erros:
            conn = get_db()
            _exec(conn,
                "UPDATE usuarios SET senha=?, trocar_senha=0 WHERE id=?",
                (generate_password_hash(nova), uid)
            )
            conn.commit()
            conn.close()
            _tokens_recuperacao.pop(token, None)
            audit('REDEFINIR_SENHA', f"uid={uid} via token")
            flash('Senha redefinida com sucesso! Faça login.', 'ok')
            return redirect(url_for('login'))

    return render_template('redefinir_senha.html', token=token, erros=erros)



@app.route('/dashboard')
def dashboard():
    if _requer_login():
        return redirect(url_for('login'))
    conn = get_db()
    mes_atual = date.today().strftime('%Y-%m')
    mes = request.args.get('mes', mes_atual)
    busca = request.args.get('busca', '').strip()
    pagina = max(1, int(request.args.get('pagina', 1)))
    por_pagina = 20

    params_base = [mes + '%']
    filtro_usuario = ""
    if session['perfil'] != 'admin':
        filtro_usuario = f"AND a.usuario_id={PH}"
        params_base.append(session['usuario_id'])

    filtro_busca = ""
    if busca:
        filtro_busca = f"AND (a.cpf LIKE {PH} OR a.nome_rf LIKE {PH})"
        params_base += [f"%{busca}%", f"%{busca}%"]

    base_where = f"WHERE a.data LIKE {PH} {filtro_usuario} {filtro_busca}"
    total_filtrado = _fetchone(conn,
        f"SELECT COUNT(*) as n FROM atendimentos a {base_where}", params_base
    )['n']

    params_mes = [mes + '%'] + ([session['usuario_id']] if session['perfil'] != 'admin' else [])
    filtro_u2 = f"AND a.usuario_id={PH}" if session['perfil'] != 'admin' else ""
    total_mes = _fetchone(conn,
        f"SELECT COUNT(*) as n FROM atendimentos a WHERE a.data LIKE {PH} {filtro_u2}",
        params_mes
    )['n']

    offset = (pagina - 1) * por_pagina
    ats = _fetchall(conn,
        f"""SELECT a.*, u.nome as entrevistador
            FROM atendimentos a JOIN usuarios u ON a.usuario_id=u.id
            {base_where} ORDER BY a.data DESC LIMIT {PH} OFFSET {PH}""",
        params_base + [por_pagina, offset]
    )

    # Total de visitas pendentes e cálculo de atrasadas (SLA)
    cfg = get_config()
    if session['perfil'] == 'admin':
        visitas_pendentes_list = _fetchall(conn, "SELECT * FROM solicitacoes_visita WHERE status='Pendente'")
    else:
        visitas_pendentes_list = _fetchall(conn,
            f"SELECT * FROM solicitacoes_visita WHERE status='Pendente' AND (solicitante_id={PH} OR responsavel_id={PH})",
            (session['usuario_id'], session['usuario_id'])
        )
    total_visitas_pendentes = len(visitas_pendentes_list)
    _, total_visitas_atrasadas = _processar_sla_visitas(visitas_pendentes_list, cfg)

    # Gráfico por Bairro / Comunidade para o Dashboard
    bairros_quant = {}
    rows_v = _fetchall(conn, "SELECT bairro FROM solicitacoes_visita WHERE criado_em LIKE ? AND bairro IS NOT NULL AND TRIM(bairro) != ''", (mes + '%',))
    for r in rows_v:
        b = r['bairro'].strip().title()
        bairros_quant[b] = bairros_quant.get(b, 0) + 1

    grafico_bairros = sorted(
        [{'nome': b, 'total': t} for b, t in bairros_quant.items()],
        key=lambda x: x['total'], reverse=True
    )[:6]

    conn.close()

    total_paginas = max(1, (total_filtrado + por_pagina - 1) // por_pagina)
    return render_template(
        'dashboard.html',
        atendimentos=ats,
        total_mes=total_mes,
        mes_atual=mes_atual,
        mes=mes,
        busca=busca,
        pagina=pagina,
        total_paginas=total_paginas,
        total_filtrado=total_filtrado,
        total_visitas_pendentes=total_visitas_pendentes,
        total_visitas_atrasadas=total_visitas_atrasadas,
        grafico_bairros=grafico_bairros,
    )

# ---------------------------------------------------------------------------
# Registrar / Editar / Excluir atendimentos
# ---------------------------------------------------------------------------

def _salvar_atendimento(conn, data, cpf, nome_rf, origem, tipos, usuario_id, at_id=None,
                        orgao_encaminhador=None, orgao_outro=None, numero_oficio=None,
                        data_encaminhamento=None, servidor_encaminhador=None,
                        motivo_encaminhamento=None, obs_encaminhamento=None,
                        situacao_encaminhamento='Atendido'):
    """Insere ou atualiza um atendimento. Retorna lista de erros ou []."""
    erros = []
    if not data:
        erros.append('Data obrigatória.')
    if not cpf:
        erros.append('CPF obrigatório.')
    elif not validar_cpf(cpf):
        erros.append('CPF inválido — verifique os dígitos.')
    if not nome_rf:
        erros.append('Nome do RF obrigatório.')
    if not origem:
        erros.append('Selecione a origem do atendimento.')
    if not tipos:
        erros.append('Selecione ao menos um tipo de atendimento.')
    if erros:
        return erros
    tipos_str = '|'.join(tipos)
    if at_id:
        _exec(conn,
            """UPDATE atendimentos SET data=?,cpf=?,nome_rf=?,origem=?,tipos=?,
               orgao_encaminhador=?,orgao_outro=?,numero_oficio=?,data_encaminhamento=?,
               servidor_encaminhador=?,motivo_encaminhamento=?,obs_encaminhamento=?,situacao_encaminhamento=?
               WHERE id=?""",
            (data, cpf, nome_rf, origem, tipos_str,
             orgao_encaminhador, orgao_outro, numero_oficio, data_encaminhamento,
             servidor_encaminhador, motivo_encaminhamento, obs_encaminhamento, situacao_encaminhamento,
             at_id)
        )
    else:
        _exec(conn,
            """INSERT INTO atendimentos (data,cpf,nome_rf,origem,tipos,usuario_id,criado_em,
               orgao_encaminhador,orgao_outro,numero_oficio,data_encaminhamento,
               servidor_encaminhador,motivo_encaminhamento,obs_encaminhamento,situacao_encaminhamento)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data, cpf, nome_rf, origem, tipos_str, usuario_id, datetime.now(_TZ_BELEM).isoformat(),
             orgao_encaminhador, orgao_outro, numero_oficio, data_encaminhamento,
             servidor_encaminhador, motivo_encaminhamento, obs_encaminhamento, situacao_encaminhamento)
        )
    conn.commit()
    return []


@app.route('/registrar', methods=['GET', 'POST'])
def registrar():
    if _requer_login():
        return redirect(url_for('login'))
    if request.method == 'POST':
        data = request.form.get('data', '').strip()
        cpf = request.form.get('cpf', '').strip()
        nome_rf = request.form.get('nome_rf', '').strip()
        origem = request.form.get('origem', '').strip()
        tipos = request.form.getlist('tipos')
        orgao_encaminhador = request.form.get('orgao_encaminhador', '').strip() or None
        orgao_outro = request.form.get('orgao_outro', '').strip() or None
        numero_oficio = request.form.get('numero_oficio', '').strip() or None
        data_encaminhamento = request.form.get('data_encaminhamento', '').strip() or None
        servidor_encaminhador = request.form.get('servidor_encaminhador', '').strip() or None
        motivo_encaminhamento = request.form.get('motivo_encaminhamento', '').strip() or None
        obs_encaminhamento = request.form.get('obs_encaminhamento', '').strip() or None
        situacao_encaminhamento = request.form.get('situacao_encaminhamento', 'Atendido').strip() or 'Atendido'

        # Bloquear SIBEC sem acesso
        if not session.get('acesso_sibec'):
            tipos = [t for t in tipos if t not in TIPOS_SIBEC]
        conn = get_db()
        erros = _salvar_atendimento(
            conn, data, cpf, nome_rf, origem, tipos, session['usuario_id'],
            orgao_encaminhador=orgao_encaminhador, orgao_outro=orgao_outro,
            numero_oficio=numero_oficio, data_encaminhamento=data_encaminhamento,
            servidor_encaminhador=servidor_encaminhador, motivo_encaminhamento=motivo_encaminhamento,
            obs_encaminhamento=obs_encaminhamento, situacao_encaminhamento=situacao_encaminhamento
        )
        conn.close()
        if erros:
            return render_template('registrar.html', erros=erros, tipos=TIPOS_ATENDIMENTO,
                                   tipos_sibec=TIPOS_SIBEC, origens=ORIGENS,
                                   orgaos_enc=ORGAOS_ENCAMINHADORES, motivos_enc=MOTIVOS_ENCAMINHAMENTO,
                                   situacoes_enc=SITUACES_ENCAMINHAMENTO,
                                   acesso_sibec=session.get('acesso_sibec'), form=request.form)
        audit('REGISTRAR_ATENDIMENTO', f"cpf={cpf} nome={nome_rf} origem={origem}")
        flash('Atendimento registrado com sucesso!', 'ok')
        return redirect(url_for('dashboard'))
    return render_template('registrar.html', erros=[], tipos=TIPOS_ATENDIMENTO,
                           tipos_sibec=TIPOS_SIBEC, origens=ORIGENS,
                           orgaos_enc=ORGAOS_ENCAMINHADORES, motivos_enc=MOTIVOS_ENCAMINHAMENTO,
                           situacoes_enc=SITUACES_ENCAMINHAMENTO,
                           acesso_sibec=session.get('acesso_sibec'), form={})


@app.route('/atendimento/<int:at_id>/editar', methods=['GET', 'POST'])
def editar_atendimento(at_id):
    if _requer_login():
        return redirect(url_for('login'))
    conn = get_db()
    at = _fetchone(conn, "SELECT * FROM atendimentos WHERE id=?", (at_id,))
    if not at or (session['perfil'] != 'admin' and at['usuario_id'] != session['usuario_id']):
        conn.close()
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        data = request.form.get('data', '').strip()
        cpf = request.form.get('cpf', '').strip()
        nome_rf = request.form.get('nome_rf', '').strip()
        origem = request.form.get('origem', '').strip()
        tipos = request.form.getlist('tipos')
        orgao_encaminhador = request.form.get('orgao_encaminhador', '').strip() or None
        orgao_outro = request.form.get('orgao_outro', '').strip() or None
        numero_oficio = request.form.get('numero_oficio', '').strip() or None
        data_encaminhamento = request.form.get('data_encaminhamento', '').strip() or None
        servidor_encaminhador = request.form.get('servidor_encaminhador', '').strip() or None
        motivo_encaminhamento = request.form.get('motivo_encaminhamento', '').strip() or None
        obs_encaminhamento = request.form.get('obs_encaminhamento', '').strip() or None
        situacao_encaminhamento = request.form.get('situacao_encaminhamento', 'Atendido').strip() or 'Atendido'

        if not session.get('acesso_sibec'):
            tipos = [t for t in tipos if t not in TIPOS_SIBEC]
        erros = _salvar_atendimento(
            conn, data, cpf, nome_rf, origem, tipos, at['usuario_id'], at_id=at_id,
            orgao_encaminhador=orgao_encaminhador, orgao_outro=orgao_outro,
            numero_oficio=numero_oficio, data_encaminhamento=data_encaminhamento,
            servidor_encaminhador=servidor_encaminhador, motivo_encaminhamento=motivo_encaminhamento,
            obs_encaminhamento=obs_encaminhamento, situacao_encaminhamento=situacao_encaminhamento
        )
        if erros:
            conn.close()
            return render_template('editar_atendimento.html', at=at, erros=erros, tipos=TIPOS_ATENDIMENTO,
                                   tipos_sibec=TIPOS_SIBEC, origens=ORIGENS,
                                   orgaos_enc=ORGAOS_ENCAMINHADORES, motivos_enc=MOTIVOS_ENCAMINHAMENTO,
                                   situacoes_enc=SITUACES_ENCAMINHAMENTO,
                                   acesso_sibec=session.get('acesso_sibec'), form=request.form)
        conn.close()
        audit('EDITAR_ATENDIMENTO', f"id={at_id} cpf={cpf}")
        flash('Atendimento atualizado com sucesso!', 'ok')
        return redirect(url_for('dashboard'))
    conn.close()
    return render_template('editar_atendimento.html', at=at, erros=[], tipos=TIPOS_ATENDIMENTO,
                           tipos_sibec=TIPOS_SIBEC, origens=ORIGENS,
                           orgaos_enc=ORGAOS_ENCAMINHADORES, motivos_enc=MOTIVOS_ENCAMINHAMENTO,
                           situacoes_enc=SITUACES_ENCAMINHAMENTO,
                           acesso_sibec=session.get('acesso_sibec'), form={})


@app.route('/atendimento/<int:at_id>/excluir', methods=['POST'])
def excluir_atendimento(at_id):
    if _requer_login():
        return redirect(url_for('login'))
    conn = get_db()
    at = _fetchone(conn, "SELECT * FROM atendimentos WHERE id=?", (at_id,))
    if at and (session['perfil'] == 'admin' or at['usuario_id'] == session['usuario_id']):
        _exec(conn, "DELETE FROM atendimentos WHERE id=?", (at_id,))
        conn.commit()
        audit('EXCLUIR_ATENDIMENTO', f"id={at_id} cpf={at['cpf']} nome={at['nome_rf']}")
        flash('Atendimento excluído.', 'ok')
    conn.close()
    return redirect(url_for('dashboard'))

# ---------------------------------------------------------------------------
# Histórico por CPF
# ---------------------------------------------------------------------------

@app.route('/api/cpf/<cpf>')
def api_cpf(cpf):
    """Retorna o nome do RF para um CPF já cadastrado (para autocomplete)."""
    if _requer_login():
        return jsonify({}), 401
    conn = get_db()
    row = _fetchone(conn,
        "SELECT nome_rf FROM atendimentos WHERE cpf=? ORDER BY criado_em DESC LIMIT 1",
        (cpf,)
    )
    conn.close()
    if row:
        return jsonify({'nome_rf': row['nome_rf']})
    return jsonify({})


@app.route('/cpf/<cpf>')
def historico_cpf(cpf):
    if _requer_login():
        return redirect(url_for('login'))
    conn = get_db()
    filtro_usuario = "" if session['perfil'] == 'admin' else f"AND a.usuario_id={PH}"
    params = [cpf] + ([session['usuario_id']] if session['perfil'] != 'admin' else [])
    ats = _fetchall(conn,
        f"""SELECT a.*, u.nome as entrevistador FROM atendimentos a
            JOIN usuarios u ON a.usuario_id=u.id
            WHERE a.cpf={PH} {filtro_usuario} ORDER BY a.data DESC""",
        params
    )
    nome_rf = ats[0]['nome_rf'] if ats else cpf
    conn.close()
    return render_template('historico_cpf.html', atendimentos=ats, cpf=cpf, nome_rf=nome_rf)


# ---------------------------------------------------------------------------
# Relatório
# ---------------------------------------------------------------------------

MESES_PT = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}


def nome_mes(mes):
    try:
        dt = datetime.strptime(mes + '-01', '%Y-%m-%d')
        return f"{MESES_PT[dt.month]} de {dt.year}"
    except ValueError:
        return mes


def dados_relatorio(mes):
    conn = get_db()
    filtro_usuario = "" if session['perfil'] == 'admin' else f"AND a.usuario_id={PH}"
    params = [mes + '%'] + ([session['usuario_id']] if session['perfil'] != 'admin' else [])
    order = "u.nome, a.data" if session['perfil'] == 'admin' else "a.data"
    atendimentos = _fetchall(conn,
        f"""SELECT a.*, u.nome as entrevistador FROM atendimentos a
            JOIN usuarios u ON a.usuario_id=u.id
            WHERE a.data LIKE {PH} {filtro_usuario} ORDER BY {order}""",
        params
    )

    quant = {t: {} for t in TIPOS_ATENDIMENTO}
    origens_quant = {o: 0 for o in ORIGENS}
    entrevistadores = []
    for at in atendimentos:
        ent = at['entrevistador']
        if ent not in entrevistadores:
            entrevistadores.append(ent)
        origens_quant[at['origem']] = origens_quant.get(at['origem'], 0) + 1
        for tipo in at['tipos'].split('|'):
            if tipo in quant:
                quant[tipo][ent] = quant[tipo].get(ent, 0) + 1

    grafico_tipos = sorted(
        [{'nome': t, 'total': sum(quant[t].values())} for t in TIPOS_ATENDIMENTO if sum(quant[t].values()) > 0],
        key=lambda x: x['total'], reverse=True
    )
    grafico_origens = [{'nome': o, 'total': t} for o, t in origens_quant.items() if t > 0]
    total_geral = sum(i['total'] for i in grafico_tipos)

    # Totais por entrevistador para gráfico
    grafico_entrevistadores = sorted(
        [{'nome': e, 'total': sum(quant[t].get(e, 0) for t in TIPOS_ATENDIMENTO)} for e in entrevistadores],
        key=lambda x: x['total'], reverse=True
    )

    # Gráfico por Bairro / Comunidade (Demanda de Atendimentos e Visitas)
    bairros_quant = {}
    rows_v = _fetchall(conn, "SELECT bairro FROM solicitacoes_visita WHERE criado_em LIKE ? AND bairro IS NOT NULL AND TRIM(bairro) != ''", (mes + '%',))
    for r in rows_v:
        b = r['bairro'].strip().title()
        bairros_quant[b] = bairros_quant.get(b, 0) + 1

    grafico_bairros = sorted(
        [{'nome': b, 'total': t} for b, t in bairros_quant.items()],
        key=lambda x: x['total'], reverse=True
    )[:8]

    # Estatísticas de Encaminhamentos por Órgão e Situação
    orgaos_quant = {}
    situacoes_enc_quant = {s: 0 for s in SITUACES_ENCAMINHAMENTO}
    for at in atendimentos:
        at_dict = dict(at)
        if at_dict.get('origem') == 'Encaminhado':
            org = at_dict.get('orgao_outro') if at_dict.get('orgao_encaminhador') == 'Outro' else at_dict.get('orgao_encaminhador')
            org = org.strip() if org else 'Não especificado'
            orgaos_quant[org] = orgaos_quant.get(org, 0) + 1

            sit = at_dict.get('situacao_encaminhamento') or 'Atendido'
            situacoes_enc_quant[sit] = situacoes_enc_quant.get(sit, 0) + 1

    grafico_orgaos = sorted(
        [{'nome': o, 'total': t} for o, t in orgaos_quant.items()],
        key=lambda x: x['total'], reverse=True
    )
    grafico_situacoes_enc = [{'nome': s, 'total': t} for s, t in situacoes_enc_quant.items() if t > 0]

    conn.close()
    return atendimentos, quant, entrevistadores, grafico_tipos, grafico_origens, total_geral, grafico_entrevistadores, grafico_bairros, grafico_orgaos, grafico_situacoes_enc


@app.route('/relatorio')
def relatorio():
    if _requer_login():
        return redirect(url_for('login'))
    mes = request.args.get('mes', date.today().strftime('%Y-%m'))
    atendimentos, quant, entrevistadores, grafico_tipos, grafico_origens, total_geral, grafico_ents, grafico_bairros, grafico_orgaos, grafico_situacoes_enc = dados_relatorio(mes)
    return render_template('relatorio.html', quant=quant, entrevistadores=entrevistadores,
                           tipos=TIPOS_ATENDIMENTO, mes=mes, atendimentos=atendimentos,
                           mes_nome=nome_mes(mes), grafico_tipos=grafico_tipos,
                           grafico_origens=grafico_origens, total_geral=total_geral,
                           grafico_entrevistadores=grafico_ents,
                           grafico_bairros=grafico_bairros,
                           grafico_orgaos=grafico_orgaos,
                           grafico_situacoes_enc=grafico_situacoes_enc)

# ---------------------------------------------------------------------------
# Exportação Excel
# ---------------------------------------------------------------------------

def usuarios_exportacao():
    conn = get_db()
    if session['perfil'] == 'admin':
        usuarios = _fetchall(conn, "SELECT * FROM usuarios ORDER BY nome")
    else:
        usuarios = _fetchall(conn, "SELECT * FROM usuarios WHERE id=?", (session['usuario_id'],))
    conn.close()
    return usuarios


def criar_excel_relatorio(mes):
    usuarios = usuarios_exportacao()
    conn = get_db()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    verde_escuro = "00542A"
    verde = "00883A"
    roxo = "783DB2"
    verde_claro = "E6F4EA"
    roxo_claro = "F3E8FF"
    azul_claro = "BDD7EE"
    cinza = "F8FAFC"
    branco = "FFFFFF"
    thin = Side(style='thin', color='CBD5E1')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    if session['perfil'] == 'admin':
        # Separar entrevistadores por unidade
        u_tome = [u for u in usuarios if (dict(u).get('unidade') or 'Tomé-Açu (Sede)') != 'Quatro Bocas']
        u_qb = [u for u in usuarios if dict(u).get('unidade') == 'Quatro Bocas']

        mes_titulo = nome_mes(mes).upper()

        # Buscar todos os atendimentos do mês
        ats_all = _fetchall(conn,
            "SELECT a.*, u.nome as entrevistador, COALESCE(u.unidade, 'Tomé-Açu (Sede)') as unidade "
            "FROM atendimentos a JOIN usuarios u ON a.usuario_id=u.id WHERE a.data LIKE ?",
            (mes + '%',)
        )

        nomes_all = [u['nome'] for u in usuarios]
        quant = {t: {n: 0 for n in nomes_all} for t in TIPOS_ATENDIMENTO}
        for at in ats_all:
            for tipo in at['tipos'].split('|'):
                if tipo in quant and at['entrevistador'] in quant[tipo]:
                    quant[tipo][at['entrevistador']] += 1

        # -------------------------------------------------------------
        # 1. ABA "Quantitativo Geral" (Consolidado Comparativo)
        # -------------------------------------------------------------
        ws_q = wb.create_sheet("Quantitativo Geral")
        ws_q.merge_cells('A1:Z1')
        ws_q['A1'] = f"QUANTITATIVO DE ATENDIMENTOS (CONSOLIDADO GERAL) - {mes_titulo}"
        ws_q['A1'].font = Font(bold=True, color=branco, size=13)
        ws_q['A1'].fill = PatternFill("solid", fgColor=verde_escuro)
        ws_q['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_q.row_dimensions[1].height = 26

        # Linha 2: Cabeçalho das Unidades
        ws_q.cell(2, 1, "UNIDADE").fill = PatternFill("solid", fgColor=verde_escuro)
        ws_q.cell(2, 1).font = Font(bold=True, color=branco)
        ws_q.cell(2, 1).border = borda
        ws_q.cell(2, 1).alignment = Alignment(horizontal='center')

        col_idx = 2
        col_tome_subtotal = None
        if u_tome:
            start_tome = col_idx
            end_tome = start_tome + len(u_tome) - 1
            for ci in range(start_tome, end_tome + 1):
                c = ws_q.cell(2, ci)
                c.fill = PatternFill("solid", fgColor=verde)
                c.border = borda
            if start_tome < end_tome:
                ws_q.merge_cells(start_row=2, start_column=start_tome, end_row=2, end_column=end_tome)
            ws_q.cell(2, start_tome, "TOMÉ-AÇU (SEDE)").font = Font(bold=True, color=branco, size=10)
            ws_q.cell(2, start_tome).alignment = Alignment(horizontal='center')
            col_idx = end_tome + 1

            c = ws_q.cell(2, col_idx, "SUBTOTAL TOMÉ")
            c.font = Font(bold=True, color=branco)
            c.fill = PatternFill("solid", fgColor=verde_escuro)
            c.border = borda
            c.alignment = Alignment(horizontal='center')
            col_tome_subtotal = col_idx
            col_idx += 1

        col_qb_subtotal = None
        if u_qb:
            start_qb = col_idx
            end_qb = start_qb + len(u_qb) - 1
            for ci in range(start_qb, end_qb + 1):
                c = ws_q.cell(2, ci)
                c.fill = PatternFill("solid", fgColor=roxo)
                c.border = borda
            if start_qb < end_qb:
                ws_q.merge_cells(start_row=2, start_column=start_qb, end_row=2, end_column=end_qb)
            ws_q.cell(2, start_qb, "QUATRO BOCAS").font = Font(bold=True, color=branco, size=10)
            ws_q.cell(2, start_qb).alignment = Alignment(horizontal='center')
            col_idx = end_qb + 1

            c = ws_q.cell(2, col_idx, "SUBTOTAL Q. BOCAS")
            c.font = Font(bold=True, color=branco)
            c.fill = PatternFill("solid", fgColor=roxo)
            c.border = borda
            c.alignment = Alignment(horizontal='center')
            col_qb_subtotal = col_idx
            col_idx += 1

        c_gt = ws_q.cell(2, col_idx, "TOTAL GERAL")
        c_gt.font = Font(bold=True, color=branco)
        c_gt.fill = PatternFill("solid", fgColor=verde_escuro)
        c_gt.border = borda
        c_gt.alignment = Alignment(horizontal='center')
        col_grand_total = col_idx
        ws_q.row_dimensions[2].height = 20

        # Linha 3: Cabeçalho com Nome dos Entrevistadores
        ws_q.cell(3, 1, "TIPO DE ATENDIMENTO").fill = PatternFill("solid", fgColor=verde_escuro)
        ws_q.cell(3, 1).font = Font(bold=True, color=branco)
        ws_q.cell(3, 1).border = borda
        ws_q.cell(3, 1).alignment = Alignment(horizontal='center')

        tome_col_map = {}
        col_curr = 2
        if u_tome:
            for u in u_tome:
                c = ws_q.cell(3, col_curr, u['nome'])
                c.font = Font(bold=True, color=branco)
                c.fill = PatternFill("solid", fgColor=verde)
                c.border = borda
                c.alignment = Alignment(horizontal='center')
                tome_col_map[u['nome']] = col_curr
                col_curr += 1
            c = ws_q.cell(3, col_curr, "SUBTOTAL TOMÉ")
            c.font = Font(bold=True, color=branco)
            c.fill = PatternFill("solid", fgColor=verde_escuro)
            c.border = borda
            c.alignment = Alignment(horizontal='center')
            col_curr += 1

        qb_col_map = {}
        if u_qb:
            for u in u_qb:
                c = ws_q.cell(3, col_curr, u['nome'])
                c.font = Font(bold=True, color=branco)
                c.fill = PatternFill("solid", fgColor=roxo)
                c.border = borda
                c.alignment = Alignment(horizontal='center')
                qb_col_map[u['nome']] = col_curr
                col_curr += 1
            c = ws_q.cell(3, col_curr, "SUBTOTAL Q. BOCAS")
            c.font = Font(bold=True, color=branco)
            c.fill = PatternFill("solid", fgColor=roxo)
            c.border = borda
            c.alignment = Alignment(horizontal='center')
            col_curr += 1

        c = ws_q.cell(3, col_curr, "TOTAL GERAL")
        c.font = Font(bold=True, color=branco)
        c.fill = PatternFill("solid", fgColor=verde_escuro)
        c.border = borda
        c.alignment = Alignment(horizontal='center')
        ws_q.row_dimensions[3].height = 20

        # Linhas por tipo de atendimento
        for ri, tipo in enumerate(TIPOS_ATENDIMENTO, 4):
            fill = PatternFill("solid", fgColor=cinza if ri % 2 == 0 else branco)
            c = ws_q.cell(ri, 1, tipo)
            c.fill = fill
            c.border = borda

            subtotal_tome = 0
            if u_tome:
                for u in u_tome:
                    v = quant[tipo].get(u['nome'], 0)
                    subtotal_tome += v
                    cc = ws_q.cell(ri, tome_col_map[u['nome']], v)
                    cc.fill = fill
                    cc.border = borda
                    cc.alignment = Alignment(horizontal='center')
                c_st = ws_q.cell(ri, col_tome_subtotal, subtotal_tome)
                c_st.font = Font(bold=True)
                c_st.fill = PatternFill("solid", fgColor=verde_claro)
                c_st.border = borda
                c_st.alignment = Alignment(horizontal='center')

            subtotal_qb = 0
            if u_qb:
                for u in u_qb:
                    v = quant[tipo].get(u['nome'], 0)
                    subtotal_qb += v
                    cc = ws_q.cell(ri, qb_col_map[u['nome']], v)
                    cc.fill = fill
                    cc.border = borda
                    cc.alignment = Alignment(horizontal='center')
                c_sqb = ws_q.cell(ri, col_qb_subtotal, subtotal_qb)
                c_sqb.font = Font(bold=True)
                c_sqb.fill = PatternFill("solid", fgColor=roxo_claro)
                c_sqb.border = borda
                c_sqb.alignment = Alignment(horizontal='center')

            tot_linha = subtotal_tome + subtotal_qb
            ct = ws_q.cell(ri, col_grand_total, tot_linha)
            ct.font = Font(bold=True)
            ct.fill = PatternFill("solid", fgColor=azul_claro)
            ct.border = borda
            ct.alignment = Alignment(horizontal='center')

        # Linha Total Rodapé
        ri_total = len(TIPOS_ATENDIMENTO) + 4
        ws_q.cell(ri_total, 1, "TOTAL").font = Font(bold=True, color=branco)
        ws_q.cell(ri_total, 1).fill = PatternFill("solid", fgColor=verde_escuro)
        ws_q.cell(ri_total, 1).border = borda

        grand_tome = 0
        if u_tome:
            for u in u_tome:
                col_tot = sum(quant[t].get(u['nome'], 0) for t in TIPOS_ATENDIMENTO)
                grand_tome += col_tot
                cc = ws_q.cell(ri_total, tome_col_map[u['nome']], col_tot)
                cc.font = Font(bold=True, color=branco)
                cc.fill = PatternFill("solid", fgColor=verde)
                cc.border = borda
                cc.alignment = Alignment(horizontal='center')
            c_st = ws_q.cell(ri_total, col_tome_subtotal, grand_tome)
            c_st.font = Font(bold=True, color=branco)
            c_st.fill = PatternFill("solid", fgColor=verde_escuro)
            c_st.border = borda
            c_st.alignment = Alignment(horizontal='center')

        grand_qb = 0
        if u_qb:
            for u in u_qb:
                col_tot = sum(quant[t].get(u['nome'], 0) for t in TIPOS_ATENDIMENTO)
                grand_qb += col_tot
                cc = ws_q.cell(ri_total, qb_col_map[u['nome']], col_tot)
                cc.font = Font(bold=True, color=branco)
                cc.fill = PatternFill("solid", fgColor=roxo)
                cc.border = borda
                cc.alignment = Alignment(horizontal='center')
            c_sqb = ws_q.cell(ri_total, col_qb_subtotal, grand_qb)
            c_sqb.font = Font(bold=True, color=branco)
            c_sqb.fill = PatternFill("solid", fgColor=roxo)
            c_sqb.border = borda
            c_sqb.alignment = Alignment(horizontal='center')

        ct = ws_q.cell(ri_total, col_grand_total, grand_tome + grand_qb)
        ct.font = Font(bold=True, color=branco)
        ct.fill = PatternFill("solid", fgColor=verde_escuro)
        ct.border = borda
        ct.alignment = Alignment(horizontal='center')

        ws_q.column_dimensions['A'].width = 35
        for ci in range(2, col_grand_total + 1):
            ws_q.column_dimensions[get_column_letter(ci)].width = 16

        # -------------------------------------------------------------
        # Helper para criar abas exclusivas por Unidade
        # -------------------------------------------------------------
        def _criar_aba_unidade(sheet_title, titulo_banner, users_list, main_color):
            ws_u = wb.create_sheet(sheet_title)
            ws_u.merge_cells('A1:Z1')
            ws_u['A1'] = f"{titulo_banner} - {mes_titulo}"
            ws_u['A1'].font = Font(bold=True, color=branco, size=13)
            ws_u['A1'].fill = PatternFill("solid", fgColor=main_color)
            ws_u['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_u.row_dimensions[1].height = 24

            headers = ["TIPO DE ATENDIMENTO"] + [u['nome'] for u in users_list] + ["TOTAL"]
            for ci, texto in enumerate(headers, 1):
                c = ws_u.cell(2, ci, texto)
                c.font = Font(bold=True, color=branco)
                c.fill = PatternFill("solid", fgColor=main_color)
                c.border = borda
                c.alignment = Alignment(horizontal='center')

            for ri, tipo in enumerate(TIPOS_ATENDIMENTO, 3):
                fill = PatternFill("solid", fgColor=cinza if ri % 2 == 0 else branco)
                c = ws_u.cell(ri, 1, tipo)
                c.fill = fill
                c.border = borda
                tot_l = 0
                for ci, u in enumerate(users_list, 2):
                    v = quant[tipo].get(u['nome'], 0)
                    tot_l += v
                    cc = ws_u.cell(ri, ci, v)
                    cc.fill = fill
                    cc.border = borda
                    cc.alignment = Alignment(horizontal='center')
                ct = ws_u.cell(ri, len(users_list) + 2, tot_l)
                ct.font = Font(bold=True)
                ct.fill = PatternFill("solid", fgColor=azul_claro)
                ct.border = borda
                ct.alignment = Alignment(horizontal='center')

            ri_tot = len(TIPOS_ATENDIMENTO) + 3
            ws_u.cell(ri_tot, 1, "TOTAL").font = Font(bold=True, color=branco)
            ws_u.cell(ri_tot, 1).fill = PatternFill("solid", fgColor=main_color)
            ws_u.cell(ri_tot, 1).border = borda
            tot_g = 0
            for ci, u in enumerate(users_list, 2):
                col_t = sum(quant[t].get(u['nome'], 0) for t in TIPOS_ATENDIMENTO)
                tot_g += col_t
                cc = ws_u.cell(ri_tot, ci, col_t)
                cc.font = Font(bold=True, color=branco)
                cc.fill = PatternFill("solid", fgColor=main_color)
                cc.border = borda
                cc.alignment = Alignment(horizontal='center')
            ct = ws_u.cell(ri_tot, len(users_list) + 2, tot_g)
            ct.font = Font(bold=True, color=branco)
            ct.fill = PatternFill("solid", fgColor=main_color)
            ct.border = borda
            ct.alignment = Alignment(horizontal='center')

            ws_u.column_dimensions['A'].width = 35
            for ci in range(2, len(users_list) + 3):
                ws_u.column_dimensions[get_column_letter(ci)].width = 16

        if u_tome:
            _criar_aba_unidade("Tomé-Açu (Sede)", "QUANTITATIVO - TOMÉ-AÇU (SEDE)", u_tome, verde_escuro)
        if u_qb:
            _criar_aba_unidade("Quatro Bocas", "QUANTITATIVO - POLO QUATRO BOCAS", u_qb, roxo)

        # -------------------------------------------------------------
        # Aba Origens
        # -------------------------------------------------------------
        ws_o = wb.create_sheet("Origens")
        ws_o.merge_cells('A1:C1')
        ws_o['A1'] = f"ATENDIMENTOS POR ORIGEM - {mes_titulo}"
        ws_o['A1'].font = Font(bold=True, color=branco, size=12)
        ws_o['A1'].fill = PatternFill("solid", fgColor=verde_escuro)
        ws_o['A1'].alignment = Alignment(horizontal='center')
        for ci, cab in enumerate(["ORIGEM", "TOTAL", "% DO TOTAL"], 1):
            c = ws_o.cell(2, ci, cab)
            c.font = Font(bold=True, color=branco)
            c.fill = PatternFill("solid", fgColor=verde_escuro)
            c.border = borda
            c.alignment = Alignment(horizontal='center')
        origens_totais = {o: 0 for o in ORIGENS}
        for at in ats_all:
            origens_totais[at['origem']] = origens_totais.get(at['origem'], 0) + 1
        total_or = sum(origens_totais.values()) or 1
        for ri, (origem, tot) in enumerate(origens_totais.items(), 3):
            fill = PatternFill("solid", fgColor=cinza if ri % 2 == 0 else branco)
            ws_o.cell(ri, 1, origem).fill = fill
            ws_o.cell(ri, 1).border = borda
            ws_o.cell(ri, 2, tot).fill = fill
            ws_o.cell(ri, 2).border = borda
            ws_o.cell(ri, 2).alignment = Alignment(horizontal='center')
            ws_o.cell(ri, 3, f"{tot/total_or*100:.1f}%").fill = fill
            ws_o.cell(ri, 3).border = borda
            ws_o.cell(ri, 3).alignment = Alignment(horizontal='center')
        ws_o.column_dimensions['A'].width = 28
        ws_o.column_dimensions['B'].width = 10
        ws_o.column_dimensions['C'].width = 12

    # -------------------------------------------------------------
    # Abas Individuais dos Entrevistadores
    # -------------------------------------------------------------
    for u in usuarios:
        unidade_u = dict(u).get('unidade') or 'Tomé-Açu (Sede)'
        ats_u = _fetchall(conn,
            "SELECT * FROM atendimentos WHERE usuario_id=? AND data LIKE ? ORDER BY data",
            (u['id'], mes + '%')
        )
        s_title = u['nome'][:24] + (" (QB)" if unidade_u == 'Quatro Bocas' else " (Sede)")
        ws = wb.create_sheet(s_title[:31])
        mes_titulo = nome_mes(mes).upper()
        ws.merge_cells('A1:F1')
        ws['A1'] = f"REGISTRO MENSAL DE ATENDIMENTOS - {mes_titulo}"
        ws['A1'].font = Font(bold=True, color=branco, size=12)
        header_color = roxo if unidade_u == 'Quatro Bocas' else verde_escuro
        ws['A1'].fill = PatternFill("solid", fgColor=header_color)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Entrevistador(a): {u['nome']}  |  Unidade: {unidade_u}"
        ws['A2'].font = Font(bold=True, size=11)
        ws['A2'].fill = PatternFill("solid", fgColor=roxo_claro if unidade_u == 'Quatro Bocas' else verde_claro)
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[2].height = 18
        for ci, cab in enumerate(['DATA', 'CPF DO RF', 'NOME DO RF', 'ORIGEM', 'TIPOS DE ATENDIMENTO', 'TOTAL'], 1):
            c = ws.cell(3, ci, cab)
            c.font = Font(bold=True, color=branco)
            c.fill = PatternFill("solid", fgColor=header_color)
            c.border = borda
            c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[3].height = 18
        for ri, at in enumerate(ats_u, 4):
            tipos_list = at['tipos'].split('|')
            fill = PatternFill("solid", fgColor=cinza if ri % 2 == 0 else branco)
            data_fmt = datetime.strptime(at['data'], '%Y-%m-%d').strftime('%d/%m/%Y')
            vals = [data_fmt, at['cpf'], at['nome_rf'], at['origem'], ', '.join(tipos_list), len(tipos_list)]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri, ci, v)
                c.fill = fill
                c.border = borda
                if ci == 6:
                    c.alignment = Alignment(horizontal='center')
                if ci == 5:
                    c.alignment = Alignment(wrap_text=True)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 8
        ws.freeze_panes = 'A4'

    conn.close()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ---------------------------------------------------------------------------
# Exportação Word / PDF (via relatorio_oficial)
# ---------------------------------------------------------------------------

def criar_word_relatorio(mes, atendimentos, quant, entrevistadores, grafico_tipos, grafico_origens, total_geral, grafico_ents=None):
    from relatorio_oficial import criar_docx_oficial
    return criar_docx_oficial(
        mes=mes,
        mes_nome=nome_mes(mes),
        atendimentos=atendimentos,
        quant=quant,
        entrevistadores=entrevistadores,
        tipos_atendimento=TIPOS_ATENDIMENTO,
        total_geral=total_geral,
        assinatura_nome=session.get('usuario_nome', 'ENTREVISTADOR'),
        assets_dir=os.path.join(BASE_DIR, 'static', 'report_assets'),
        config=get_config(),
    )


def criar_pdf_relatorio(mes, atendimentos, quant, entrevistadores, grafico_tipos, grafico_origens, total_geral, grafico_ents=None):
    from relatorio_oficial import criar_pdf_oficial
    return criar_pdf_oficial(
        mes=mes,
        mes_nome=nome_mes(mes),
        atendimentos=atendimentos,
        quant=quant,
        entrevistadores=entrevistadores,
        tipos_atendimento=TIPOS_ATENDIMENTO,
        total_geral=total_geral,
        assinatura_nome=session.get('usuario_nome', 'ENTREVISTADOR'),
        assets_dir=os.path.join(BASE_DIR, 'static', 'report_assets'),
        config=get_config(),
    )


def criar_pdf_registro(mes, atendimentos, quant=None, total_geral=0):
    from relatorio_oficial import criar_pdf_registro_detalhado
    return criar_pdf_registro_detalhado(
        mes=mes,
        mes_nome=nome_mes(mes),
        atendimentos=atendimentos,
        quant=quant,
        total_geral=total_geral,
        assinatura_nome=session.get('usuario_nome', 'ENTREVISTADOR'),
        assets_dir=os.path.join(BASE_DIR, 'static', 'report_assets'),
    )


def criar_word_registro(mes, atendimentos, quant=None, total_geral=0):
    from relatorio_oficial import criar_docx_registro_detalhado
    return criar_docx_registro_detalhado(
        mes=mes,
        mes_nome=nome_mes(mes),
        atendimentos=atendimentos,
        quant=quant,
        total_geral=total_geral,
        assinatura_nome=session.get('usuario_nome', 'ENTREVISTADOR'),
        assets_dir=os.path.join(BASE_DIR, 'static', 'report_assets'),
    )


@app.route('/exportar/planilha')
def exportar_planilha():
    if _requer_login():
        return redirect(url_for('login'))
    mes = request.args.get('mes', date.today().strftime('%Y-%m'))
    output = criar_excel_relatorio(mes)
    return send_file(output, as_attachment=True,
                     download_name=f"Atendimentos_CadUnico_{mes}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/exportar/relatorio')
def exportar_relatorio():
    if _requer_login():
        return redirect(url_for('login'))
    mes = request.args.get('mes', date.today().strftime('%Y-%m'))
    formato = request.args.get('formato', 'excel').lower()
    atendimentos, quant, entrevistadores, grafico_tipos, grafico_origens, total_geral, grafico_ents, grafico_bairros, grafico_orgaos, grafico_situacoes_enc = dados_relatorio(mes)

    if formato == 'word':
        try:
            output = criar_word_relatorio(mes, atendimentos, quant, entrevistadores,
                                          grafico_tipos, grafico_origens, total_geral)
        except RuntimeError as erro:
            return str(erro), 500
        return send_file(output, as_attachment=True,
                         download_name=f"Relatorio_CadUnico_{mes}.docx",
                         mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    if formato == 'pdf':
        try:
            output = criar_pdf_relatorio(mes, atendimentos, quant, entrevistadores,
                                         grafico_tipos, grafico_origens, total_geral)
        except RuntimeError as erro:
            return str(erro), 500
        return send_file(output, as_attachment=True,
                         download_name=f"Relatorio_CadUnico_{mes}.pdf",
                         mimetype='application/pdf')
    output = criar_excel_relatorio(mes)
    return send_file(output, as_attachment=True,
                     download_name=f"Relatorio_CadUnico_{mes}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/exportar/registro')
def exportar_registro():
    if _requer_login():
        return redirect(url_for('login'))
    mes = request.args.get('mes', date.today().strftime('%Y-%m'))
    formato = request.args.get('formato', 'pdf').lower()
    atendimentos, quant, entrevistadores, grafico_tipos, grafico_origens, total_geral, grafico_ents, grafico_bairros, grafico_orgaos, grafico_situacoes_enc = dados_relatorio(mes)

    if formato == 'word':
        try:
            output = criar_word_registro(mes, atendimentos, quant=quant, total_geral=total_geral)
        except RuntimeError as erro:
            return str(erro), 500
        return send_file(output, as_attachment=True,
                         download_name=f"Registro_Detalhado_{mes}.docx",
                         mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    # default: pdf
    try:
        output = criar_pdf_registro(mes, atendimentos, quant=quant, total_geral=total_geral)
    except RuntimeError as erro:
        return str(erro), 500
    return send_file(output, as_attachment=True,
                     download_name=f"Registro_Detalhado_{mes}.pdf",
                     mimetype='application/pdf')

# ---------------------------------------------------------------------------
# Admin: gestão de usuários
# ---------------------------------------------------------------------------

@app.route('/admin/usuarios')
def admin_usuarios():
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    conn = get_db()
    usuarios = _fetchall(conn, "SELECT * FROM usuarios ORDER BY nome")
    conn.close()
    return render_template('admin_usuarios.html', usuarios=usuarios)


@app.route('/admin/usuarios/novo', methods=['GET', 'POST'])
def novo_usuario():
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    erros = []
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        login_ = request.form.get('login', '').strip()
        senha = request.form.get('senha', '').strip()
        perfil = request.form.get('perfil', 'entrevistador')
        unidade = request.form.get('unidade', 'Tomé-Açu (Sede)').strip()
        acesso_sibec = 1 if request.form.get('acesso_sibec') else 0
        email = request.form.get('email', '').strip() or None
        telefone = request.form.get('telefone', '').strip() or None
        if not nome:
            erros.append('Nome obrigatório.')
        if not login_:
            erros.append('Login obrigatório.')
        if not senha:
            erros.append('Senha obrigatória.')
        elif len(senha) < 6:
            erros.append('A senha deve ter ao menos 6 caracteres.')
        if not erros:
            try:
                conn = get_db()
                _exec(conn,
                    "INSERT INTO usuarios (nome,login,senha,perfil,acesso_sibec,trocar_senha,email,telefone,unidade) VALUES (?,?,?,?,?,?,?,?,?)",
                    (nome, login_, generate_password_hash(senha), perfil, acesso_sibec, 1, email, telefone, unidade)
                )
                conn.commit()
                conn.close()
                audit('CRIAR_USUARIO', f"login={login_}")
                flash(f'Usuário {nome} criado. O usuário deverá trocar a senha no primeiro acesso.', 'ok')
                return redirect(url_for('admin_usuarios'))
            except sqlite3.IntegrityError:
                erros.append('Login já existe.')
    return render_template('novo_usuario.html', erros=erros, form=request.form)


@app.route('/admin/usuarios/<int:uid>/editar', methods=['GET', 'POST'])
def editar_usuario(uid):
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    conn = get_db()
    u = _fetchone(conn, "SELECT * FROM usuarios WHERE id=?", (uid,))
    erros = []
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        login_ = request.form.get('login', '').strip()
        nova_senha = request.form.get('senha', '').strip()
        perfil = request.form.get('perfil', 'entrevistador')
        unidade = request.form.get('unidade', 'Tomé-Açu (Sede)').strip()
        acesso_sibec = 1 if request.form.get('acesso_sibec') else 0
        email = request.form.get('email', '').strip() or None
        telefone = request.form.get('telefone', '').strip() or None
        if not nome:
            erros.append('Nome obrigatório.')
        if nova_senha and len(nova_senha) < 6:
            erros.append('A nova senha deve ter ao menos 6 caracteres.')
        if not erros:
            if nova_senha:
                _exec(conn,
                    "UPDATE usuarios SET nome=?,login=?,senha=?,perfil=?,acesso_sibec=?,trocar_senha=0,email=?,telefone=?,unidade=? WHERE id=?",
                    (nome, login_, generate_password_hash(nova_senha), perfil, acesso_sibec, email, telefone, unidade, uid)
                )
            else:
                _exec(conn,
                    "UPDATE usuarios SET nome=?,login=?,perfil=?,acesso_sibec=?,email=?,telefone=?,unidade=? WHERE id=?",
                    (nome, login_, perfil, acesso_sibec, email, telefone, unidade, uid)
                )
            conn.commit()
            conn.close()
            audit('EDITAR_USUARIO', f"uid={uid} login={login_}")
            flash('Usuário atualizado com sucesso.', 'ok')
            return redirect(url_for('admin_usuarios'))
    conn.close()
    return render_template('editar_usuario.html', u=u, erros=erros)


@app.route('/admin/usuarios/<int:uid>/excluir', methods=['POST'])
def excluir_usuario(uid):
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    if uid != session['usuario_id']:
        conn = get_db()
        u = _fetchone(conn, "SELECT login FROM usuarios WHERE id=?", (uid,))
        _exec(conn, "DELETE FROM usuarios WHERE id=?", (uid,))
        conn.commit()
        conn.close()
        audit('EXCLUIR_USUARIO', f"uid={uid} login={u['login'] if u else '?'}")
        flash('Usuário excluído.', 'ok')
    return redirect(url_for('admin_usuarios'))


# ---------------------------------------------------------------------------
# Admin: log de auditoria
# ---------------------------------------------------------------------------

@app.route('/admin/zerar-numeracao-vd', methods=['POST'])
def zerar_numeracao_vd():
    if session.get('perfil') != 'admin':
        return redirect(url_for('painel_visitas'))
    conn = get_db()
    _exec(conn, "DELETE FROM visita_contadores")
    _exec(conn, "UPDATE solicitacoes_visita SET numero_vd = NULL")
    conn.commit()
    conn.close()
    audit('NUMERACAO_VD_ZERADA', 'contador zerado pelo admin')
    flash('Numeração VD zerada. A próxima visita receberá VD-2026-000001.', 'ok')
    return redirect(url_for('painel_visitas'))


@app.route('/admin/auditoria')
def auditoria():
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    conn = get_db()
    registros = _fetchall(conn,
        "SELECT * FROM audit_log ORDER BY criado_em DESC LIMIT 200"
    )
    conn.close()
    return render_template('auditoria.html', registros=registros)


@app.route('/admin/central-backup')
def central_backup():
    if _requer_login() or session.get('perfil') != 'admin':
        flash('Acesso negado.', 'erro')
        return redirect(url_for('dashboard'))

    conn = get_db()
    total_usuarios = _fetchone(conn, "SELECT COUNT(*) as n FROM usuarios")['n']
    total_atendimentos = _fetchone(conn, "SELECT COUNT(*) as n FROM atendimentos")['n']
    total_visitas = _fetchone(conn, "SELECT COUNT(*) as n FROM solicitacoes_visita")['n']
    total_audit = _fetchone(conn, "SELECT COUNT(*) as n FROM audit_log")['n']
    conn.close()

    tamanho_db_kb = 0
    db_path = os.path.join(app.root_path, 'cadunico.db')
    if not os.path.exists(db_path):
        db_path = os.path.join(BASE_DIR, 'cadunico.db')
    if os.path.exists(db_path):
        tamanho_db_kb = round(os.path.getsize(db_path) / 1024, 1)

    agora_str = datetime.now(_TZ_BELEM).strftime('%d/%m/%Y às %H:%M:%S')

    stats = {
        'total_usuarios': total_usuarios,
        'total_atendimentos': total_atendimentos,
        'total_visitas': total_visitas,
        'total_audit': total_audit,
        'tamanho_db_kb': tamanho_db_kb,
        'tamanho_db_mb': round(tamanho_db_kb / 1024, 2),
        'agora_str': agora_str,
    }

    return render_template('central_backup.html', stats=stats)


@app.route('/admin/backup')
def admin_backup():
    if _requer_login() or session.get('perfil') != 'admin':
        flash('Acesso negado.', 'erro')
        return redirect(url_for('dashboard'))

    import zipfile
    import json

    buf = io.BytesIO()
    hoje_str = datetime.now(_TZ_BELEM).strftime('%Y-%m-%d_%H%M%S')
    zip_filename = f"Backup_CadUnico_COMPLETO_{hoje_str}.zip"

    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        conn = get_db()

        if not _is_pg():
            db_path = os.path.join(app.root_path, 'cadunico.db')
            if not os.path.exists(db_path):
                db_path = os.path.join(BASE_DIR, 'cadunico.db')
            if os.path.exists(db_path):
                zf.write(db_path, arcname='banco_dados/cadunico.db')

        tabelas = [
            'usuarios', 'atendimentos', 'solicitacoes_visita',
            'config_relatorio', 'audit_log', 'visita_contadores', 'visita_fotos'
        ]
        manifest_counts = {}
        for tab in tabelas:
            try:
                rows = [dict(r) for r in _fetchall(conn, f"SELECT * FROM {tab}")]
                manifest_counts[tab] = len(rows)
                zf.writestr(f"tabelas_json/dados_{tab}.json", json.dumps(rows, indent=2, ensure_ascii=False))
            except Exception as e:
                manifest_counts[tab] = f"Erro: {str(e)}"
        conn.close()

        upload_dirs = [
            os.path.join(app.root_path, 'uploads'),
            os.path.join(BASE_DIR, 'uploads'),
            os.path.join(app.root_path, 'static', 'uploads'),
        ]
        for u_dir in upload_dirs:
            if os.path.exists(u_dir):
                for root, dirs, files in os.walk(u_dir):
                    for f in files:
                        full_p = os.path.join(root, f)
                        rel_p = os.path.relpath(full_p, u_dir)
                        zf.write(full_p, arcname=os.path.join('arquivos_anexos', rel_p))

        manifest = {
            'sistema': 'Sistema de Gestão do Cadastro Único e Programa Bolsa Família',
            'municipio': 'Tomé-Açu / PA',
            'versao_sistema': '2.5.0',
            'data_geracao': datetime.now(_TZ_BELEM).isoformat(),
            'fuso_horario': 'America/Belem (UTC-3)',
            'tabelas_backup': manifest_counts,
        }
        zf.writestr('manifesto_backup.json', json.dumps(manifest, indent=2, ensure_ascii=False))

    buf.seek(0)
    audit('BACKUP_SISTEMA_ZIP', f"arquivo={zip_filename}")
    return send_file(
        buf,
        as_attachment=True,
        download_name=zip_filename,
        mimetype='application/zip'
    )


@app.route('/admin/backup/excel')
def backup_excel():
    if _requer_login() or session.get('perfil') != 'admin':
        flash('Acesso negado.', 'erro')
        return redirect(url_for('dashboard'))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    conn = get_db()

    ws_at = wb.active
    ws_at.title = "Atendimentos"
    ats = [dict(r) for r in _fetchall(conn, "SELECT a.*, u.nome as entrevistador FROM atendimentos a JOIN usuarios u ON a.usuario_id=u.id ORDER BY a.id DESC")]
    headers_at = ['ID', 'Data', 'CPF', 'Nome do RF', 'Origem', 'Tipos', 'Entrevistador', 'Criado em', 'Órgão Encaminhador', 'Situação']
    ws_at.append(headers_at)
    for at in ats:
        ws_at.append([
            at.get('id'), at.get('data'), at.get('cpf'), at.get('nome_rf'),
            at.get('origem'), at.get('tipos'), at.get('entrevistador'),
            at.get('criado_em'), at.get('orgao_encaminhador') or '—', at.get('situacao_encaminhamento') or 'Atendido'
        ])

    ws_v = wb.create_sheet("Visitas Domiciliares")
    visitas = [dict(r) for r in _fetchall(conn, "SELECT sv.*, sol.nome as solicitante_nome, res.nome as responsavel_nome FROM solicitacoes_visita sv LEFT JOIN usuarios sol ON sol.id=sv.solicitante_id LEFT JOIN usuarios res ON res.id=sv.responsavel_id ORDER BY sv.id DESC")]
    headers_v = ['ID', 'Nº VD', 'CPF RF', 'Nome RF', 'Logradouro', 'Número', 'Bairro', 'Motivo', 'Status', 'Data Realizada', 'Solicitante', 'Responsável', 'Parecer Técnico']
    ws_v.append(headers_v)
    for v in visitas:
        ws_v.append([
            v.get('id'), v.get('numero_vd') or f"#{v.get('id')}", v.get('cpf_rf'), v.get('nome_rf'),
            v.get('logradouro'), v.get('numero'), v.get('bairro'), v.get('motivo'),
            v.get('status'), v.get('data_realizada') or '—', v.get('solicitante_nome') or '—',
            v.get('responsavel_nome') or '—', v.get('parecer_tecnico_txt') or '—'
        ])

    ws_u = wb.create_sheet("Usuários")
    users = [dict(r) for r in _fetchall(conn, "SELECT id, nome, login, perfil, unidade, email, telefone, acesso_sibec, tentativas_login FROM usuarios ORDER BY id")]
    headers_u = ['ID', 'Nome', 'Login', 'Perfil', 'Unidade', 'E-mail', 'Telefone', 'Acesso SIBEC', 'Falhas Login']
    ws_u.append(headers_u)
    for u in users:
        ws_u.append([
            u.get('id'), u.get('nome'), u.get('login'), u.get('perfil'),
            u.get('unidade'), u.get('email') or '—', u.get('telefone') or '—',
            'Sim' if u.get('acesso_sibec') else 'Não', u.get('tentativas_login', 0)
        ])

    ws_aud = wb.create_sheet("Histórico Auditoria")
    audits = [dict(r) for r in _fetchall(conn, "SELECT id, usuario_nome, acao, detalhe, ip, criado_em FROM audit_log ORDER BY id DESC LIMIT 1000")]
    headers_aud = ['ID', 'Usuário', 'Ação', 'Detalhes', 'IP', 'Data/Hora']
    ws_aud.append(headers_aud)
    for a in audits:
        ws_aud.append([
            a.get('id'), a.get('usuario_nome') or '—', a.get('acao'),
            a.get('detalhe') or '—', a.get('ip') or '—', a.get('criado_em')
        ])

    conn.close()

    fill_hdr = PatternFill("solid", fgColor="00542A")
    font_hdr = Font(bold=True, color="FFFFFF")
    for ws in [ws_at, ws_v, ws_u, ws_aud]:
        for cell in ws[1]:
            cell.fill = fill_hdr
            cell.font = font_hdr
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    hoje_str = datetime.now(_TZ_BELEM).strftime('%Y-%m-%d')
    xlsx_filename = f"Backup_Dados_CadUnico_{hoje_str}.xlsx"

    audit('BACKUP_SISTEMA_EXCEL', f"arquivo={xlsx_filename}")
    return send_file(
        buf,
        as_attachment=True,
        download_name=xlsx_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/admin/config-relatorio', methods=['GET', 'POST'])
def config_relatorio():
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        campos = [
            'coordenadora', 'setor_nome', 'endereco', 'email_setor',
            'territorio', 'texto_identificacao', 'texto_apresentacao',
            'rodape', 'municipio', 'email_admin_notificacao',
            'smtp_host', 'smtp_port', 'smtp_user', 'smtp_pass',
        ]
        for key, val in request.form.items():
            if key in campos or key.startswith('prazo_visita_'):
                val_strip = val.strip()
                set_config(key, val_strip)
        
        file = request.files.get('imagem_cabecalho')
        if file and file.filename:
            ext = os.path.splitext(file.filename)[1].lower()
            if ext in ('.png', '.jpg', '.jpeg', '.webp'):
                target_dir = os.path.join(app.root_path, 'static', 'report_assets')
                os.makedirs(target_dir, exist_ok=True)
                target_path = os.path.join(target_dir, 'image3.png')
                file.save(target_path)
            else:
                flash('Formato de imagem inválido. Use PNG, JPG ou JPEG.', 'erro')
                return redirect(url_for('config_relatorio'))

        audit('CONFIG_RELATORIO', 'configurações salvas')
        flash('Configurações salvas com sucesso!', 'ok')
        return redirect(url_for('config_relatorio'))
    cfg = get_config()
    return render_template('config_relatorio.html', cfg=cfg)


@app.route('/admin/testar-email', methods=['POST'])
def testar_email_smtp():
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    cfg = get_config()
    destino = cfg.get('email_admin_notificacao') or cfg.get('email_setor') or 'setascadastrounico@gmail.com'
    assunto = "🧪 Teste de Notificação SMTP - Sistema CadÚnico"
    html = f"""
    <div style="font-family: Arial, sans-serif; padding:20px; border:1px solid #00883A; border-radius:8px; background:#FFFFFF;">
      <h3 style="color:#00542A; margin-top:0;">✅ Teste de E-mail Bem-Sucedido!</h3>
      <p>As configurações de e-mail SMTP do Sistema CadÚnico estão funcionando corretamente.</p>
      <p><strong>E-mail de Notificação do Administrador:</strong> <code>{destino}</code></p>
    </div>
    """
    sucesso, msg = _enviar_email_smtp(destino, assunto, html)
    if sucesso:
        flash(f"E-mail de teste enviado com sucesso para {destino}!", "ok")
    else:
        flash(f"Falha ao enviar e-mail de teste: {msg}", "erro")
    return redirect(url_for('config_relatorio'))


@app.route('/admin/usuarios/<int:uid>/resetar-tentativas', methods=['POST'])
def resetar_tentativas_usuario(uid):
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    conn = get_db()
    _exec(conn, "UPDATE usuarios SET tentativas_login=0, trocar_senha=1 WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    audit('RESETAR_TENTATIVAS_SENHA', f"uid={uid}")
    flash("Tentativas incorretas zeradas! O usuário solicitará nova senha no próximo acesso.", "ok")
    return redirect(url_for('admin_usuarios'))


@app.route('/admin/config-visita', methods=['GET', 'POST'])
def config_visita():
    if session.get('perfil') != 'admin':
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        campos = [
            'visita_titulo_doc', 'visita_subtitulo_doc',
            'visita_orientacao_texto', 'visita_assinatura_1',
            'visita_assinatura_2', 'visita_rodape_txt',
        ]
        for c in campos:
            val = request.form.get(c, '').strip()
            if val:
                set_config(c, val)
        audit('CONFIG_VISITA', 'modelo impresso da solicitação de visita atualizado')
        flash('Modelo impresso da solicitação de visita atualizado com sucesso!', 'ok')
        return redirect(url_for('config_visita'))
    cfg = get_config()
    return render_template('config_visita.html', cfg=cfg)


# ---------------------------------------------------------------------------
# Controle de Solicitações de Visita
# ---------------------------------------------------------------------------

@app.route('/visitas', methods=['GET'])
def painel_visitas():
    if _requer_login():
        return redirect(url_for('login'))

    conn = get_db()
    uid = session['usuario_id']
    perfil = session.get('perfil')

    # ── Parâmetros de query ─────────────────────────────────────────────────
    status_filtro = request.args.get('status', '').strip()
    data_ini      = request.args.get('data_ini', '').strip()
    data_fim      = request.args.get('data_fim', '').strip()
    busca         = request.args.get('busca', '').strip()
    zona_filtro   = request.args.get('zona', '').strip()
    bairro_filtro = request.args.get('bairro', '').strip()
    pagina        = max(1, int(request.args.get('pagina', 1)))
    por_pagina    = 20

    # ── Filtro de acesso por perfil ─────────────────────────────────────────
    if perfil != 'admin':
        filtro_acesso  = f"AND (sv.solicitante_id = {PH} OR sv.responsavel_id = {PH})"
        params_acesso  = [uid, uid]
    else:
        filtro_acesso  = ""
        params_acesso  = []

    # ── Filtros opcionais ───────────────────────────────────────────────────
    filtro_status = ""
    params_status = []
    if status_filtro:
        filtro_status = f"AND sv.status = {PH}"
        params_status = [status_filtro]

    filtro_data = ""
    params_data = []
    if data_ini and data_fim:
        filtro_data = f"AND sv.criado_em BETWEEN {PH} AND {PH}"
        params_data = [data_ini, data_fim + 'T23:59:59']
    elif data_ini:
        filtro_data = f"AND sv.criado_em >= {PH}"
        params_data = [data_ini]
    elif data_fim:
        filtro_data = f"AND sv.criado_em <= {PH}"
        params_data = [data_fim + 'T23:59:59']

    filtro_busca = ""
    params_busca = []
    if busca:
        filtro_busca = f"AND (sv.cpf_rf LIKE {PH} OR LOWER(sv.nome_rf) LIKE LOWER({PH}))"
        params_busca = [f"%{busca}%", f"%{busca}%"]

    filtro_zona = ""
    params_zona = []
    if zona_filtro:
        filtro_zona = f"AND sv.zona = {PH}"
        params_zona = [zona_filtro]

    filtro_bairro = ""
    params_bairro = []
    if bairro_filtro:
        filtro_bairro = f"AND LOWER(sv.bairro) LIKE LOWER({PH})"
        params_bairro = [f"%{bairro_filtro}%"]

    # ── Monta WHERE completo ────────────────────────────────────────────────
    where = f"WHERE 1=1 {filtro_acesso} {filtro_status} {filtro_data} {filtro_busca} {filtro_zona} {filtro_bairro}"
    params_where = params_acesso + params_status + params_data + params_busca + params_zona + params_bairro

    # ── Contagem total filtrada ─────────────────────────────────────────────
    total_filtrado = _fetchone(conn,
        f"""SELECT COUNT(*) as n
            FROM solicitacoes_visita sv
            {where}""",
        params_where
    )['n']

    # ── Paginação ───────────────────────────────────────────────────────────
    offset        = (pagina - 1) * por_pagina
    total_paginas = max(1, (total_filtrado + por_pagina - 1) // por_pagina)

    cfg = get_config()

    # ── Buscar todas as pendentes para calcular total de atrasadas em todo o acervo
    where_acesso = f"WHERE 1=1 {filtro_acesso}"
    visitas_pendentes_all = _fetchall(conn,
        f"""SELECT sv.* FROM solicitacoes_visita sv {where_acesso} AND sv.status='Pendente'""",
        params_acesso
    )
    _, total_visitas_atrasadas = _processar_sla_visitas(visitas_pendentes_all, cfg)

    # ── Se o filtro de status for "Atrasada", busca todas as pendentes e filtra por SLA
    if status_filtro == 'Atrasada':
        visitas_raw = _fetchall(conn,
            f"""SELECT sv.*, sol.nome AS solicitante_nome, res.nome AS responsavel_nome
                FROM solicitacoes_visita sv
                JOIN usuarios sol ON sv.solicitante_id = sol.id
                LEFT JOIN usuarios res ON sv.responsavel_id = res.id
                {where_acesso} AND sv.status='Pendente'
                ORDER BY sv.criado_em DESC""",
            params_acesso
        )
        visitas_proc, _ = _processar_sla_visitas(visitas_raw, cfg)
        visitas = [v for v in visitas_proc if v['atrasada']]
        total_filtrado = len(visitas)
        total_paginas = max(1, (total_filtrado + por_pagina - 1) // por_pagina)
        visitas = visitas[(pagina - 1) * por_pagina : pagina * por_pagina]
    else:
        visitas_raw = _fetchall(conn,
            f"""SELECT sv.*,
                       sol.nome  AS solicitante_nome,
                       res.nome  AS responsavel_nome
                FROM solicitacoes_visita sv
                JOIN usuarios sol ON sv.solicitante_id = sol.id
                LEFT JOIN usuarios res ON sv.responsavel_id = res.id
                {where}
                ORDER BY sv.criado_em DESC
                LIMIT {PH} OFFSET {PH}""",
            params_where + [por_pagina, offset]
        )
        visitas, _ = _processar_sla_visitas(visitas_raw, cfg)

    # ── Contadores por status ───────────────────────────────────────────────
    contadores_rows = _fetchall(conn,
        f"""SELECT sv.status, COUNT(*) as total
            FROM solicitacoes_visita sv
            {where_acesso}
            GROUP BY sv.status""",
        params_acesso
    )
    contadores = {'Pendente': 0, 'Em Andamento': 0, 'Realizada': 0, 'Cancelada': 0}
    for row in contadores_rows:
        if row['status'] in contadores:
            contadores[row['status']] = row['total']

    # ── Lista de usuários (apenas para admin) ───────────────────────────────
    usuarios = []
    if perfil == 'admin':
        usuarios = _fetchall(conn, "SELECT id, nome FROM usuarios ORDER BY nome")

    conn.close()

    filtros = {
        'status':   status_filtro,
        'data_ini': data_ini,
        'data_fim': data_fim,
        'busca':    busca,
        'zona':     zona_filtro,
        'bairro':   bairro_filtro,
    }

    return render_template(
        'painel_visitas.html',
        visitas=visitas,
        pagina=pagina,
        total_paginas=total_paginas,
        total_filtrado=total_filtrado,
        filtros=filtros,
        contadores=contadores,
        total_visitas_atrasadas=total_visitas_atrasadas,
        usuarios=usuarios,
    )


@app.route('/visitas/<int:visita_id>', methods=['GET'])
def detalhe_visita(visita_id):
    if _requer_login():
        return redirect(url_for('login'))

    conn = get_db()
    uid = session['usuario_id']
    perfil = session.get('perfil')

    # Busca a solicitação; se não existir, flash de erro e redireciona
    visita = _fetchone(conn,
        "SELECT * FROM solicitacoes_visita WHERE id=?",
        (visita_id,)
    )
    if not visita:
        conn.close()
        flash('Solicitação não encontrada.', 'erro')
        return redirect(url_for('painel_visitas'))

    # Verifica permissão: entrevistador só acessa se for solicitante ou responsável
    if perfil != 'admin':
        if visita['solicitante_id'] != uid and visita['responsavel_id'] != uid:
            conn.close()
            flash('Acesso negado.', 'erro')
            return redirect(url_for('painel_visitas'))

    # Busca dados do solicitante e do responsável
    solicitante = _fetchone(conn,
        "SELECT * FROM usuarios WHERE id=?",
        (visita['solicitante_id'],)
    )
    responsavel = None
    if visita['responsavel_id']:
        responsavel = _fetchone(conn,
            "SELECT * FROM usuarios WHERE id=?",
            (visita['responsavel_id'],)
        )

    # Calcula flags de permissão
    STATUS_TERMINAIS = ('Realizada', 'Cancelada', 'Não Localizada')
    status_nao_terminal = visita['status'] not in STATUS_TERMINAIS
    tem_permissao = (perfil == 'admin') or (
        visita['solicitante_id'] == uid or visita['responsavel_id'] == uid
    )
    pode_editar = status_nao_terminal and tem_permissao
    pode_excluir = perfil == 'admin'
    pode_registrar_resultado = status_nao_terminal and tem_permissao

    # Busca fotos da residência
    fotos = _fetchall(conn,
        "SELECT * FROM visita_fotos WHERE solicitacao_id=? ORDER BY criado_em",
        (visita_id,)
    )

    # Busca histórico de auditoria para esta solicitação
    historico = _fetchall(conn,
        f"SELECT * FROM audit_log WHERE detalhe LIKE {PH} ORDER BY criado_em ASC",
        (f"%id={visita_id}%",)
    )

    conn.close()

    return render_template(
        'detalhe_visita.html',
        visita=visita,
        solicitante=solicitante,
        responsavel=responsavel,
        pode_editar=pode_editar,
        pode_excluir=pode_excluir,
        pode_registrar_resultado=pode_registrar_resultado,
        fotos=fotos,
        historico=historico,
    )


@app.route('/visitas/<int:visita_id>/parecer-tecnico', methods=['POST'])
def emitir_parecer_tecnico_visita(visita_id):
    if _requer_login():
        return redirect(url_for('login'))

    conn = get_db()
    visita = _fetchone(conn, "SELECT * FROM solicitacoes_visita WHERE id=?", (visita_id,))
    if not visita:
        conn.close()
        flash('Solicitação não encontrada.', 'erro')
        return redirect(url_for('painel_visitas'))

    parecer_txt = request.form.get('parecer_tecnico_txt', '').strip()
    if not parecer_txt:
        conn.close()
        flash('Por favor, informe o texto do Parecer Técnico Assistencial.', 'erro')
        return redirect(url_for('detalhe_visita', visita_id=visita_id))

    parecer_url = dict(visita).get('parecer_as_url')
    parecer_nome = dict(visita).get('parecer_as_nome')

    arquivo_parecer = request.files.get('parecer_as_file')
    if arquivo_parecer and arquivo_parecer.filename:
        err = _validar_parecer(arquivo_parecer)
        if err:
            conn.close()
            flash(err, 'erro')
            return redirect(url_for('detalhe_visita', visita_id=visita_id))
        url_p, nome_p = _upload_anexo(arquivo_parecer, pasta='visitas_pareceres')
        if url_p:
            parecer_url = url_p
            parecer_nome = nome_p

    agora = datetime.now(_TZ_BELEM).isoformat()
    hoje_str = date.today().isoformat()

    _exec(conn,
        """UPDATE solicitacoes_visita
           SET parecer_tecnico_txt=?, parecer_as_url=?, parecer_as_nome=?,
               status='Realizada', data_realizada=?, responsavel_id=?,
               atualizado_em=?
           WHERE id=?""",
        (parecer_txt, parecer_url, parecer_nome, hoje_str, session['usuario_id'], agora, visita_id)
    )
    conn.commit()
    conn.close()

    audit('PARECER_TECNICO_EMITIDO', f"id={visita_id} por {session['usuario_nome']}")
    flash('Parecer Técnico Assistencial registrado e visita concluída com sucesso!', 'ok')
    return redirect(url_for('detalhe_visita', visita_id=visita_id))


@app.route('/visitas/<int:visita_id>/editar', methods=['GET', 'POST'])
def editar_visita(visita_id):
    if _requer_login():
        return redirect(url_for('login'))

    conn = get_db()
    uid = session['usuario_id']
    perfil = session.get('perfil')

    STATUS_TERMINAIS = ('Realizada', 'Cancelada', 'Não Localizada')

    # Busca a solicitação; se não existir, flash de erro e redireciona
    visita = _fetchone(conn,
        "SELECT * FROM solicitacoes_visita WHERE id=?",
        (visita_id,)
    )
    if not visita:
        conn.close()
        flash('Solicitação não encontrada.', 'erro')
        return redirect(url_for('painel_visitas'))

    # Verifica permissão: entrevistador só edita se for solicitante ou responsável
    if perfil != 'admin':
        if visita['solicitante_id'] != uid and visita['responsavel_id'] != uid:
            conn.close()
            flash('Acesso negado.', 'erro')
            return redirect(url_for('painel_visitas'))

    # Admin pode atribuir responsável; entrevistador não
    usuarios = []
    if perfil == 'admin':
        usuarios = _fetchall(conn, "SELECT id, nome FROM usuarios ORDER BY nome")

    if request.method == 'POST':
        # Verifica status terminal antes de processar POST
        if visita['status'] in STATUS_TERMINAIS:
            conn.close()
            flash('Esta solicitação não pode ser editada pois já foi finalizada.', 'erro')
            return redirect(url_for('detalhe_visita', visita_id=visita_id))

        nome_rf       = request.form.get('nome_rf', '').strip()
        logradouro    = request.form.get('logradouro', '').strip()
        numero        = request.form.get('numero', '').strip()
        complemento   = request.form.get('complemento', '').strip() or None
        bairro        = request.form.get('bairro', '').strip()
        referencia    = request.form.get('referencia', '').strip() or None
        zona          = request.form.get('zona', 'Urbana').strip()
        motivo        = request.form.get('motivo', '').strip()
        observacoes   = request.form.get('observacoes', '').strip() or None
        telefone1     = request.form.get('telefone1', '').strip() or None
        telefone2     = request.form.get('telefone2', '').strip() or None

        # Apenas admin pode alterar o responsável
        if perfil == 'admin':
            responsavel_id = request.form.get('responsavel_id', '').strip() or None
        else:
            responsavel_id = visita['responsavel_id']

        # Upload de novo anexo (se fornecido)
        anexo_url  = visita['anexo_url']
        anexo_nome = visita['anexo_nome']
        arquivo = request.files.get('anexo')
        if arquivo and arquivo.filename:
            url_nova, nome_orig = _upload_anexo(arquivo, pasta='visitas')
            if url_nova:
                anexo_url  = url_nova
                anexo_nome = nome_orig

        # Upload de novo parecer AS (se fornecido)
        parecer_url  = visita['parecer_as_url']
        parecer_nome = visita['parecer_as_nome']
        arquivo_parecer = request.files.get('parecer_as')
        if arquivo_parecer and arquivo_parecer.filename:
            err = _validar_parecer(arquivo_parecer)
            if err:
                erros.append(err)
            else:
                url_p, nome_p = _upload_anexo(arquivo_parecer, pasta='visitas_pareceres')
                if url_p:
                    parecer_url  = url_p
                    parecer_nome = nome_p

        form = {
            'nome_rf': nome_rf,
            'logradouro': logradouro,
            'numero': numero,
            'complemento': complemento or '',
            'bairro': bairro,
            'referencia': referencia or '',
            'zona': zona,
            'motivo': motivo,
            'responsavel_id': responsavel_id,
            'observacoes': observacoes or '',
        }

        erros = []
        if not nome_rf:
            erros.append('Nome do RF é obrigatório.')
        if not logradouro:
            erros.append('Logradouro é obrigatório.')
        if not numero:
            erros.append('Número é obrigatório.')
        if not bairro:
            erros.append('Bairro é obrigatório.')
        if not motivo:
            erros.append('Motivo é obrigatório.')

        if erros:
            conn.close()
            return render_template('editar_visita.html',
                                   visita=visita, erros=erros,
                                   form=form, usuarios=usuarios)

        agora = datetime.now(_TZ_BELEM).isoformat()
        _exec(conn,
            """UPDATE solicitacoes_visita
               SET nome_rf=?, logradouro=?, numero=?, complemento=?, bairro=?,
                   referencia=?, zona=?, motivo=?,
                   responsavel_id=?, observacoes=?, anexo_url=?, anexo_nome=?,
                   parecer_as_url=?, parecer_as_nome=?,
                   telefone1=?, telefone2=?,
                   atualizado_em=?
               WHERE id=?""",
            (nome_rf, logradouro, numero, complemento, bairro,
             referencia, zona, motivo,
             responsavel_id, observacoes, anexo_url, anexo_nome,
             parecer_url, parecer_nome,
             telefone1, telefone2,
             agora, visita_id)
        )
        conn.commit()
        conn.close()
        audit('VISITA_EDITADA', f"id={visita_id} editado por {session['usuario_nome']}")
        flash('Solicitação atualizada com sucesso!', 'ok')
        return redirect(url_for('detalhe_visita', visita_id=visita_id))

    # GET — verifica status terminal antes de exibir o formulário
    if visita['status'] in STATUS_TERMINAIS:
        conn.close()
        flash('Esta solicitação não pode ser editada pois já foi finalizada.', 'erro')
        return redirect(url_for('detalhe_visita', visita_id=visita_id))

    # Pré-preenche o formulário com os dados atuais
    form = {
        'nome_rf':        visita['nome_rf'],
        'logradouro':     visita['logradouro'] or '',
        'numero':         visita['numero'] or '',
        'complemento':    visita['complemento'] or '',
        'bairro':         visita['bairro'] or '',
        'referencia':     visita['referencia'] or '',
        'zona':           visita['zona'] or 'Urbana',
        'motivo':         visita['motivo'],
        'responsavel_id': visita['responsavel_id'],
        'observacoes':    visita['observacoes'] or '',
    }

    conn.close()
    return render_template('editar_visita.html',
                           visita=visita, erros=[], form=form, usuarios=usuarios)


@app.route('/visitas/nova', methods=['GET', 'POST'])
def nova_visita():
    if _requer_login():
        return redirect(url_for('login'))

    conn = get_db()
    erros = []
    form = {}

    # Admin pode atribuir um responsável na criação
    usuarios = []
    if session.get('perfil') == 'admin':
        usuarios = _fetchall(conn, "SELECT id, nome FROM usuarios ORDER BY nome")

    if request.method == 'POST':
        cpf_rf      = request.form.get('cpf_rf', '').strip()
        nome_rf     = request.form.get('nome_rf', '').strip()
        logradouro  = request.form.get('logradouro', '').strip()
        numero      = request.form.get('numero', '').strip()
        complemento = request.form.get('complemento', '').strip() or None
        bairro      = request.form.get('bairro', '').strip()
        referencia  = request.form.get('referencia', '').strip() or None
        zona        = request.form.get('zona', 'Urbana').strip()
        motivo      = request.form.get('motivo', '').strip()
        motivo_especificado = request.form.get('motivo_especificado', '').strip() or None
        responsavel_id = request.form.get('responsavel_id', '').strip() or None

        # Entrevistador sempre se autoatribui como responsável
        if session.get('perfil') != 'admin':
            responsavel_id = str(session['usuario_id'])
        observacoes = request.form.get('observacoes', '').strip() or None
        telefone1   = request.form.get('telefone1', '').strip() or None
        telefone2   = request.form.get('telefone2', '').strip() or None

        # Se motivo for "Outro", usa o campo especificado
        motivo_final = motivo_especificado if motivo == 'Outro' and motivo_especificado else motivo

        # Upload de anexo (opcional)
        anexo_url  = None
        anexo_nome = None
        arquivo = request.files.get('anexo')
        if arquivo and arquivo.filename:
            anexo_url, anexo_nome = _upload_anexo(arquivo, pasta='visitas')

        form = {
            'cpf_rf': cpf_rf,
            'nome_rf': nome_rf,
            'logradouro': logradouro,
            'numero': numero,
            'complemento': complemento or '',
            'bairro': bairro,
            'referencia': referencia or '',
            'zona': zona,
            'motivo': motivo,
            'responsavel_id': responsavel_id,
            'observacoes': observacoes or '',
            'telefone1': telefone1 or '',
            'telefone2': telefone2 or '',
        }

        # Validações
        if not cpf_rf:
            erros.append('CPF do RF é obrigatório.')
        elif not validar_cpf(cpf_rf):
            erros.append('CPF inválido — verifique os dígitos.')

        if not nome_rf:
            erros.append('Nome do RF é obrigatório.')
        if not logradouro:
            erros.append('Logradouro é obrigatório.')
        if not numero:
            erros.append('Número é obrigatório.')
        if not bairro:
            erros.append('Bairro é obrigatório.')
        if not motivo:
            erros.append('Motivo é obrigatório.')

        ano_belem = datetime.now(_TZ_BELEM).year

        if not erros:
            agora = datetime.now(_TZ_BELEM).isoformat()
            try:
                numero_vd = _gerar_numero_vd(conn, ano_belem)
                if _USE_PG:
                    cur = _exec(conn,
                        """INSERT INTO solicitacoes_visita
                            (cpf_rf, nome_rf, logradouro, numero, complemento, bairro,
                             referencia, zona, motivo, status, solicitante_id,
                             responsavel_id, observacoes, anexo_url, anexo_nome,
                             telefone1, telefone2, criado_em, atualizado_em, numero_vd)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'Pendente',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                           RETURNING id""",
                        (cpf_rf, nome_rf, logradouro, numero, complemento, bairro,
                         referencia, zona, motivo, session['usuario_id'],
                         responsavel_id, observacoes, anexo_url, anexo_nome,
                         telefone1, telefone2, agora, agora, numero_vd)
                    )
                    novo_id = cur.fetchone()['id']
                else:
                    _exec(conn,
                        """INSERT INTO solicitacoes_visita
                            (cpf_rf, nome_rf, logradouro, numero, complemento, bairro,
                             referencia, zona, motivo, status, solicitante_id,
                             responsavel_id, observacoes, anexo_url, anexo_nome,
                             telefone1, telefone2, criado_em, atualizado_em, numero_vd)
                           VALUES (?,?,?,?,?,?,?,?,?,'Pendente',?,?,?,?,?,?,?,?,?,?)""",
                        (cpf_rf, nome_rf, logradouro, numero, complemento, bairro,
                         referencia, zona, motivo, session['usuario_id'],
                         responsavel_id, observacoes, anexo_url, anexo_nome,
                         telefone1, telefone2, agora, agora, numero_vd)
                    )
                    novo_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            except ValueError as e:
                if str(e) == 'limite_anual':
                    conn.close()
                    flash('Limite de solicitações para o ano atingido. Contate o administrador.', 'erro')
                    return render_template('nova_visita.html', erros=[], form=form, usuarios=usuarios)
                conn.rollback()
                conn.close()
                flash('Erro ao gerar número da solicitação. Tente novamente.', 'erro')
                return render_template('nova_visita.html', erros=[], form=form, usuarios=usuarios)
            except Exception:
                conn.rollback()
                conn.close()
                flash('Erro ao gerar número da solicitação. Tente novamente.', 'erro')
                return render_template('nova_visita.html', erros=[], form=form, usuarios=usuarios)

            conn.commit()
            conn.close()
            audit('VISITA_CRIADA', f"cpf={cpf_rf} nome={nome_rf} id={novo_id}")
            flash('Solicitação de visita registrada com sucesso!', 'ok')
            return redirect(url_for('painel_visitas'))

        conn.close()
        return render_template('nova_visita.html', erros=erros, form=form, usuarios=usuarios)

    # GET — pré-preencher nome_rf a partir do CPF se fornecido na query string
    cpf_qs = request.args.get('cpf', '').strip()
    if cpf_qs:
        atendimento_recente = _fetchone(conn,
            "SELECT nome_rf FROM atendimentos WHERE cpf=? ORDER BY criado_em DESC LIMIT 1",
            (cpf_qs,)
        )
        form['cpf_rf'] = cpf_qs
        if atendimento_recente:
            form['nome_rf'] = atendimento_recente['nome_rf']

    conn.close()
    return render_template('nova_visita.html', erros=erros, form=form, usuarios=usuarios)


@app.route('/visitas/<int:visita_id>/status', methods=['POST'])
def atualizar_status_visita(visita_id):
    if _requer_login():
        return redirect(url_for('login'))

    conn = get_db()
    uid = session['usuario_id']
    perfil = session.get('perfil')

    STATUSES_VALIDOS = {'Pendente', 'Em Andamento', 'Realizada', 'Cancelada', 'Não Localizada'}
    STATUS_TERMINAIS = {'Realizada', 'Cancelada', 'Não Localizada'}

    visita = _fetchone(conn, "SELECT * FROM solicitacoes_visita WHERE id=?", (visita_id,))
    if not visita:
        conn.close()
        flash('Solicitação não encontrada.', 'erro')
        return redirect(url_for('painel_visitas'))

    if perfil != 'admin':
        if visita['solicitante_id'] != uid and visita['responsavel_id'] != uid:
            conn.close()
            flash('Acesso negado.', 'erro')
            return redirect(url_for('painel_visitas'))

    status_anterior = visita['status']

    if status_anterior in STATUS_TERMINAIS:
        conn.close()
        flash('Esta solicitação já foi finalizada e não pode ter o status alterado.', 'erro')
        return redirect(url_for('detalhe_visita', visita_id=visita_id))

    novo_status = request.form.get('novo_status', '').strip()

    if novo_status not in ('Cancelada', 'Não Localizada'):
        conn.close()
        flash('Selecione uma opção válida: Família Não Localizada ou Cancelar Solicitação.', 'erro')
        return redirect(url_for('detalhe_visita', visita_id=visita_id))

    motivo_cancelamento = request.form.get('motivo_cancelamento', '').strip() or None

    if not motivo_cancelamento:
        conn.close()
        flash('Por favor, informe a justificativa / motivo.', 'erro')
        return redirect(url_for('detalhe_visita', visita_id=visita_id))

    agora = datetime.now(_TZ_BELEM).isoformat()

    _exec(conn,
        """UPDATE solicitacoes_visita
           SET status=?, motivo_cancelamento=?, atualizado_em=?
           WHERE id=?""",
        (novo_status, motivo_cancelamento, agora, visita_id)
    )
    conn.commit()
    conn.close()

    audit('VISITA_STATUS_ATUALIZADO', f"id={visita_id} status={status_anterior}→{novo_status}")
    if novo_status == 'Não Localizada':
        flash('Solicitação registrada como Família Não Localizada com sucesso!', 'ok')
    else:
        flash('Solicitação de visita cancelada com sucesso!', 'ok')
    return redirect(url_for('detalhe_visita', visita_id=visita_id))


@app.route('/visitas/<int:visita_id>/excluir', methods=['POST'])
def excluir_visita(visita_id):
    if _requer_login():
        return redirect(url_for('login'))

    if session.get('perfil') != 'admin':
        return redirect(url_for('painel_visitas'))

    conn = get_db()
    visita = _fetchone(conn, "SELECT * FROM solicitacoes_visita WHERE id=?", (visita_id,))
    if visita:
        _exec(conn, "DELETE FROM solicitacoes_visita WHERE id=?", (visita_id,))
        conn.commit()
        conn.close()
        audit('VISITA_EXCLUIDA', f"id={visita_id} cpf={visita['cpf_rf']}")
        flash('Solicitação excluída com sucesso.', 'ok')
    else:
        conn.close()
        flash('Solicitação não encontrada.', 'erro')

    return redirect(url_for('painel_visitas'))


# ---------------------------------------------------------------------------
# PDF da solicitação de visita
# ---------------------------------------------------------------------------

@app.route('/visitas/<int:visita_id>/pdf', methods=['GET'])
def pdf_visita(visita_id):
    if _requer_login():
        return redirect(url_for('login'))

    conn = get_db()
    visita = _fetchone(conn, "SELECT * FROM solicitacoes_visita WHERE id=?", (visita_id,))
    if not visita:
        conn.close()
        abort(404)
    conn.close()

    pdf_bytes, numero_vd = gerar_pdf_visita(visita_id)
    if pdf_bytes is None:
        abort(404)
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f"{numero_vd}.pdf",
    )


# ---------------------------------------------------------------------------
# Registro de resultado da visita
# ---------------------------------------------------------------------------

@app.route('/visitas/<int:visita_id>/resultado', methods=['GET', 'POST'])
def resultado_visita(visita_id):
    if _requer_login():
        return redirect(url_for('login'))

    conn   = get_db()
    uid    = session['usuario_id']
    perfil = session.get('perfil')

    visita = _fetchone(conn, "SELECT * FROM solicitacoes_visita WHERE id=?", (visita_id,))
    if not visita:
        conn.close()
        flash('Solicitação não encontrada.', 'erro')
        return redirect(url_for('painel_visitas'))

    STATUS_TERMINAIS = ('Realizada', 'Cancelada')
    if visita['status'] in STATUS_TERMINAIS:
        conn.close()
        flash('Esta solicitação já foi finalizada.', 'erro')
        return redirect(url_for('detalhe_visita', visita_id=visita_id))

    tem_permissao = (perfil == 'admin' or
                     visita['solicitante_id'] == uid or
                     visita['responsavel_id'] == uid)
    if not tem_permissao:
        conn.close()
        flash('Acesso negado.', 'erro')
        return redirect(url_for('painel_visitas'))

    erros = []
    hoje  = date.today().isoformat()

    if request.method == 'POST':
        data_realizada = request.form.get('data_realizada', '').strip()
        observacoes    = request.form.get('observacoes', '').strip() or None
        arquivo_parecer = request.files.get('parecer_as')

        if not data_realizada:
            erros.append('Informe a data de realização da visita.')
        elif data_realizada > hoje:
            erros.append('A data de realização não pode ser uma data futura.')

        if not erros:
            agora = datetime.now(_TZ_BELEM).isoformat()
            parecer_url  = visita['parecer_as_url']
            parecer_nome = visita['parecer_as_nome']
            aviso_parecer = None

            if arquivo_parecer and arquivo_parecer.filename:
                err = _validar_parecer(arquivo_parecer)
                if err:
                    erros.append(err)
                else:
                    url_nova, nome_orig = _upload_anexo(arquivo_parecer, pasta='visitas_pareceres')
                    if url_nova:
                        parecer_url  = url_nova
                        parecer_nome = nome_orig
                    else:
                        aviso_parecer = ('Resultado registrado, mas o parecer não pôde ser '
                                         'anexado. Tente enviá-lo novamente.')

        if not erros:
            cpf_rf  = visita['cpf_rf']
            nome_rf = visita['nome_rf']
            try:
                if _USE_PG:
                    cur = _exec(conn,
                        """INSERT INTO atendimentos
                            (data, cpf, nome_rf, origem, tipos, usuario_id, criado_em)
                           VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                        (data_realizada, cpf_rf, nome_rf,
                         'Visita Domiciliar', 'Visita Domiciliar', uid, agora)
                    )
                    atendimento_id = cur.fetchone()['id']
                else:
                    _exec(conn,
                        """INSERT INTO atendimentos
                            (data, cpf, nome_rf, origem, tipos, usuario_id, criado_em)
                           VALUES (?,?,?,?,?,?,?)""",
                        (data_realizada, cpf_rf, nome_rf,
                         'Visita Domiciliar', 'Visita Domiciliar', uid, agora)
                    )
                    atendimento_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

                _exec(conn,
                    """UPDATE solicitacoes_visita
                       SET status='Realizada', data_realizada=?, observacoes=?,
                           parecer_as_url=?, parecer_as_nome=?,
                           atendimento_id=?, atualizado_em=?
                       WHERE id=?""",
                    (data_realizada, observacoes, parecer_url, parecer_nome,
                     atendimento_id, agora, visita_id)
                )
                conn.commit()
            except Exception:
                conn.rollback()
                conn.close()
                flash('Erro ao registrar resultado. Tente novamente.', 'erro')
                return redirect(url_for('detalhe_visita', visita_id=visita_id))

            conn.close()
            audit('VISITA_RESULTADO', f"id={visita_id} data={data_realizada}")
            if aviso_parecer:
                flash(aviso_parecer, 'aviso')
            flash('Resultado registrado com sucesso!', 'ok')
            return redirect(url_for('detalhe_visita', visita_id=visita_id))

    conn.close()
    return render_template('resultado_visita.html',
                           visita=visita, erros=erros, hoje=hoje)


# ---------------------------------------------------------------------------
# Upload de múltiplas fotos da residência
# ---------------------------------------------------------------------------

@app.route('/visitas/<int:visita_id>/fotos', methods=['POST'])
def upload_fotos_visita(visita_id):
    if _requer_login():
        return redirect(url_for('login'))

    conn   = get_db()
    uid    = session['usuario_id']
    perfil = session.get('perfil')

    visita = _fetchone(conn, "SELECT * FROM solicitacoes_visita WHERE id=?", (visita_id,))
    if not visita:
        conn.close()
        flash('Solicitação não encontrada.', 'erro')
        return redirect(url_for('painel_visitas'))

    STATUS_TERMINAIS = ('Realizada', 'Cancelada')
    if visita['status'] in STATUS_TERMINAIS:
        conn.close()
        flash('Não é possível adicionar fotos a uma solicitação finalizada.', 'erro')
        return redirect(url_for('detalhe_visita', visita_id=visita_id))

    tem_permissao = (perfil == 'admin' or
                     visita['solicitante_id'] == uid or
                     visita['responsavel_id'] == uid)
    if not tem_permissao:
        conn.close()
        flash('Acesso negado.', 'erro')
        return redirect(url_for('painel_visitas'))

    row_count = _fetchone(conn,
        "SELECT COUNT(*) as n FROM visita_fotos WHERE solicitacao_id=?", (visita_id,))
    fotos_existentes = row_count['n'] if row_count else 0

    arquivos  = request.files.getlist('fotos')
    sucessos  = 0
    erros_msg = []
    agora     = datetime.now().isoformat()

    for arq in arquivos:
        if not arq or not arq.filename:
            continue
        if fotos_existentes + sucessos >= 10:
            erros_msg.append('Limite de 10 fotos por solicitação atingido.')
            break
        err = _validar_foto(arq)
        if err:
            erros_msg.append(err)
            continue
        url_nova, nome_orig = _upload_anexo(arq, pasta='visitas_fotos')
        if url_nova:
            _exec(conn,
                "INSERT INTO visita_fotos (solicitacao_id, url, nome_arquivo, criado_em) VALUES (?,?,?,?)",
                (visita_id, url_nova, nome_orig, agora)
            )
            sucessos += 1
        else:
            erros_msg.append(f"Falha ao enviar '{arq.filename}'.")

    if sucessos > 0:
        conn.commit()
        audit('VISITA_FOTOS', f"id={visita_id} fotos={sucessos}")

    conn.close()

    if sucessos > 0:
        flash(f'{sucessos} foto(s) adicionada(s) com sucesso.', 'ok')
    if erros_msg:
        flash('Erros: ' + ' | '.join(dict.fromkeys(erros_msg)), 'erro')

    return redirect(url_for('detalhe_visita', visita_id=visita_id))


# ---------------------------------------------------------------------------
# Histórico de visitas por família (CPF)
# ---------------------------------------------------------------------------

@app.route('/visitas/familia/<cpf>', methods=['GET'])
def historico_familia(cpf):
    if _requer_login():
        return redirect(url_for('login'))

    cpf_digits = ''.join(c for c in cpf if c.isdigit())
    if not validar_cpf(cpf_digits):
        return render_template('historico_familia.html',
                               visitas=[], cpf=cpf, nome_rf=None,
                               erro='CPF inválido.'), 400

    conn   = get_db()
    uid    = session['usuario_id']
    perfil = session.get('perfil')

    if perfil == 'admin':
        visitas = _fetchall(conn, """
            SELECT sv.*, u.nome as responsavel_nome
            FROM solicitacoes_visita sv
            LEFT JOIN usuarios u ON u.id = sv.responsavel_id
            WHERE sv.cpf_rf = ?
            ORDER BY sv.criado_em DESC
        """, (cpf_digits,))
    else:
        visitas = _fetchall(conn, """
            SELECT sv.*, u.nome as responsavel_nome
            FROM solicitacoes_visita sv
            LEFT JOIN usuarios u ON u.id = sv.responsavel_id
            WHERE sv.cpf_rf = ? AND (sv.solicitante_id = ? OR sv.responsavel_id = ?)
            ORDER BY sv.criado_em DESC
        """, (cpf_digits, uid, uid))

    nome_rf = visitas[0]['nome_rf'] if visitas else None
    conn.close()
    return render_template('historico_familia.html',
                           visitas=visitas, cpf=cpf_digits,
                           nome_rf=nome_rf, erro=None)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=False)
else:
    # Executado pelo gunicorn em produção
    init_db()
